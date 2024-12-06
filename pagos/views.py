# Imports estándar de Python
import os
import io
import csv
import logging
import warnings
from datetime import datetime
from calendar import month_name
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
from django.db.models.functions import TruncMonth

# Imports de terceros
import pandas as pd
import openpyxl
from fuzzywuzzy import fuzz, process
import pytz
import calendar
# Imports de Django
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.db import transaction
from django.db.models import F, Q, Value, Sum, Count, Avg, DecimalField, CharField
from django.db.models.functions import Coalesce, TruncMonth

from django.http import Http404, HttpResponse, JsonResponse, FileResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils.text import slugify
from django.utils.translation import gettext as _

# Imports de Django Rest Framework
from rest_framework import viewsets, status, filters
from rest_framework.decorators import api_view
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.authentication import TokenAuthentication

# Imports de la aplicación 'pagos'
from .forms import (
    CapacityForm,
    CapacityFormSet,
    CustomAuthenticationForm,
    CustomUserCreationForm,
    ExcelUploadForm,
    LPUExcelUploadForm,
    LPUForm,
    NuevaTablaPagosForm,
    SLAReportForm,
    GenerarTablaSLAMesForm
)
from .models import (
    Capacity,
    LPU,
    NuevaTablaPagos,
    TablaDePagoCapacity,
    SLAReport,
    ZonaAdjudicacion,
    ZonaOperativa,
    PrecioEspecialidad,
    Especialidad
)
from .serializers import (
    CapacitySerializer,
    NuevaTablaPagosSerializer,
    TablaDePagoCapacitySerializer,
    SLAReportSerializer
)
from .utils import mapear_especialidad, obtener_proyecto, clasificar_especialidad, limpiar_texto_especialidad

# Configuración del logger
logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')


# Vistas de autenticación y navegación

# Función que maneja la vista principal de la aplicación
@login_required
def index(request):
    # Verifica si el usuario está autenticado
    if request.user.is_authenticated:
        # Si el usuario está autenticado, lo redirige a la vista 'menu'
        return redirect('menu')
    # Si el usuario no está autenticado, lo redirige a la página de inicio de sesión 'login'
    return redirect('login')

from django.contrib.auth import login, authenticate, logout
from django.shortcuts import render, redirect
from django.contrib import messages
from pagos.forms import CustomUserCreationForm, CustomAuthenticationForm  # Asegúrate de importar tus formularios personalizados
from django.contrib.auth.decorators import login_required

# Vista para el registro de nuevos usuarios
def register(request):
    # Verifica si el método de la solicitud es POST, lo que indica que el usuario ha enviado el formulario
    if request.method == 'POST':
        # Se crea una instancia del formulario personalizado de creación de usuario con los datos enviados por el usuario
        form = CustomUserCreationForm(request.POST)
        # Verifica si el formulario es válido (pasa todas las validaciones definidas)
        if form.is_valid():
            # Si el formulario es válido, se guarda el nuevo usuario en la base de datos
            user = form.save()
            # Autentica al nuevo usuario automáticamente después del registro
            login(request, user)
            # Redirige al usuario autenticado al menú principal
            return redirect('menu')
    else:
        # Si el método no es POST (es decir, es una solicitud GET), se crea un formulario vacío para ser rellenado por el usuario
        form = CustomUserCreationForm()
    # Renderiza la plantilla 'register.html' pasándole el formulario para que el usuario lo complete
    return render(request, 'pagos/register.html', {'form': form})

# Vista personalizada para el inicio de sesión
def login_view(request):
    # Verifica si la solicitud es de tipo POST (el formulario ha sido enviado)
    if request.method == 'POST':
        # Crea una instancia del formulario de autenticación personalizado con los datos enviados
        form = CustomAuthenticationForm(data=request.POST)
        # Verifica si el formulario es válido (es decir, que ha pasado todas las validaciones)
        if form.is_valid():
            # Obtiene el 'username' y 'password' de los datos validados del formulario
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            # Autentica al usuario utilizando el nombre de usuario y la contraseña
            user = authenticate(username=username, password=password)
            # Si el usuario es encontrado (es decir, las credenciales son correctas)
            if user is not None:
                # Inicia la sesión para el usuario autenticado
                login(request, user)
                # Redirige al usuario al menú principal
                return redirect('menu')
            # Si no se encuentra un usuario con las credenciales proporcionadas, muestra un mensaje de error
            messages.error(request, 'Por favor, ingrese un RUT y contraseña correctos.')
        else:
            # Si el formulario no es válido, muestra un mensaje de error indicando que hay errores en el formulario
            messages.error(request, 'Por favor corrija los errores en el formulario.')
    else:
        # Si el método no es POST (es decir, es una solicitud GET), se muestra un formulario vacío para el inicio de sesión
        form = CustomAuthenticationForm()
    # Renderiza la plantilla 'login.html' pasándole el formulario de autenticación para que el usuario lo complete
    return render(request, 'pagos/login.html', {'form': form})

# Vista para cerrar la sesión del usuario
@login_required
def logout_view(request):
    # Llama a la función logout de Django, que cierra la sesión del usuario actual
    logout(request)
    # Redirige al usuario a la página de inicio de sesión ('login') después de cerrar la sesión
    return redirect('login')



# Vista para el menú principal, accesible solo para usuarios autenticados
@login_required
def menu(request):
    # Renderiza la plantilla 'menu.html' y pasa el objeto 'user' (el usuario autenticado) como contexto a la plantilla
    return render(request, 'menu.html', {'user': request.user})


@login_required
def cargar_nueva_tabla_pagos(request):
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            
            # Verificar que el archivo subido sea un archivo Excel
            if not excel_file.name.endswith('.xlsx'):
                messages.error(request, 'Por favor, suba un archivo Excel con extensión .xlsx.')
                return render(request, 'pagos/cargar_nueva_tabla_pagos.html', {'form': form})
            
            try:
                # Leer el archivo Excel en un DataFrame de pandas, suprimiendo advertencias
                with warnings.catch_warnings():
                    warnings.simplefilter("ignore")
                    df = pd.read_excel(excel_file, engine='openpyxl').rename(columns=str.lower).rename(columns=lambda x: x.strip())
                
                # Renombrar columnas con nombres alternativos
                column_mapping = {
                    'desvio_porcentaje': 'desvio_porcentaje'
                }
                df.rename(columns=column_mapping, inplace=True)
                
                # Asegurarse de que las columnas del Excel coinciden con las del modelo
                expected_columns = ['sociedad', 'material', 'especialidad', 'zona_adjudicacion', 'zona_operacional', 'nuevo_precio', 'desvio_porcentaje', 'area']
                missing_columns = [column for column in expected_columns if column not in df.columns]
                if missing_columns:
                    messages.error(request, f'El archivo Excel no contiene las columnas esperadas. Faltan: {missing_columns}')
                    messages.info(request, f'Las columnas encontradas en el archivo son: {list(df.columns)}')
                    return render(request, 'pagos/cargar_nueva_tabla_pagos.html', {'form': form})

                # Iterar sobre las filas del DataFrame y crear o actualizar los registros en la base de datos
                for _, row in df.iterrows():
                    sociedad = str(row['sociedad']).zfill(4)  # Asegurar que tenga 4 dígitos
                    material = row['material']
                    especialidad = row['especialidad']
                    zona_adjudicacion = row['zona_adjudicacion']
                    zona_operacional = row['zona_operacional']
                    nuevo_precio = row['nuevo_precio']
                    desvio_porcentaje = row['desvio_porcentaje']
                    area = row['area']
                    anio = datetime.now().year  # Asignar el año actual

                    if pd.isna(sociedad) or pd.isna(material) or pd.isna(especialidad):
                        messages.warning(request, 'Se encontró una fila con datos nulos. Fila omitida.')
                        continue

                    # Crear o actualizar el registro en la base de datos
                    NuevaTablaPagos.objects.update_or_create(
                        sociedad=sociedad,
                        material=material,
                        especialidad=especialidad,
                        zona_adjudicacion=zona_adjudicacion,
                        zona_operacional=zona_operacional,
                        area=area,
                        defaults={
                            'nuevo_precio': nuevo_precio,
                            'desvio_porcentaje': desvio_porcentaje,
                            'anio': anio
                        }
                    )

                messages.success(request, 'El archivo Excel se ha cargado correctamente.')
                return render(request, 'pagos/cargar_nueva_tabla_pagos.html', {'form': form})

            except Exception as e:
                messages.error(request, f'Error al procesar el archivo: {str(e)}')
                return render(request, 'pagos/cargar_nueva_tabla_pagos.html', {'form': form})
    else:
        form = ExcelUploadForm()
    
    return render(request, 'pagos/cargar_nueva_tabla_pagos.html', {'form': form})



@login_required
def cargar_excel(request):
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            
            # Asegurarse de que el archivo tiene la extensión correcta
            if not excel_file.name.endswith('.xlsx') and not excel_file.name.endswith('.xls'):
                messages.error(request, 'El archivo debe tener una extensión .xlsx o .xls')
                return redirect('cargar_excel')

            # Guardar el archivo Excel y procesar los datos
            try:
                data_path = settings.BASE_DIR / 'gestion_pagos' / 'pagos' / 'data'
                data_path.mkdir(parents=True, exist_ok=True)
                
                file_path = data_path / f'CAPACITY_SERVICE_{datetime.now().strftime("%Y-%m-%d_%H-%M-%S")}.xlsx'
                with open(file_path, 'wb') as destination:
                    for chunk in excel_file.chunks():
                        destination.write(chunk)

                # Leer el archivo Excel
                df = pd.read_excel(file_path)

                # Renombrar columnas para que sean más fáciles de manejar
                df.rename(columns={
                    'ESPECIALIDAD USUARIO': 'ESPECIALIDAD_USUARIO',
                    'RUT TECNICO': 'RUT_TECNICO',
                    'ZONA OPERACIONAL': 'ZONA_OPERACIONAL',
                    'ZONA ADJUDICACION': 'ZONA_ADJUDICACION',
                    'TOTAL HH': 'TOTAL_HH',
                    'TOTAL HH + PHI': 'TOTAL_HH_PHI',
                    'ESPECIALIDAD TEAM': 'ESPECIALIDAD_TEAM'
                }, inplace=True)

                # Validar columnas
                expected_columns = {
                    'FECHA_EXTRACT_REGISTRO': ['FECHA_EXTRACT_REGISTRO', 'fecha_extract_registro'],
                    'NOMBRE_TECNICO': ['NOMBRE_TECNICO', 'nombre_tecnico'],
                    'NOMBRE_TEAM': ['NOMBRE_TEAM', 'nombre_team'],
                    'ESPECIALIDAD_USUARIO': ['ESPECIALIDAD_USUARIO', 'especialidad usuario'],
                    'AREA': ['AREA', 'area'],
                    'RUT_TECNICO': ['RUT_TECNICO', 'rut tecnico'],
                    'ZONA_OPERACIONAL': ['ZONA_OPERACIONAL', 'zona operacional'],
                    'ZONA_ADJUDICACION': ['ZONA_ADJUDICACION', 'zona adjudicacion'],
                    'ESPECIALIDAD_TEAM': ['ESPECIALIDAD_TEAM', 'especialidad team', 'especialidad equipo', 'especialidad'],
                    'CAPACIDAD': ['CAPACIDAD', 'capacidad'],
                    'TOTAL_HH': ['TOTAL_HH', 'total hh'],
                    'PHI': ['PHI', 'phi'],
                    'TOTAL_HH_PHI': ['TOTAL_HH_PHI', 'total hh + phi'],
                }

                missing_columns = [col for col, alt_names in expected_columns.items() if not any(alt in df.columns for alt in alt_names)]

                if missing_columns:
                    messages.error(request, f'El archivo no contiene las columnas necesarias: {", ".join(missing_columns)}')
                    messages.info(request, f'Las columnas encontradas en el archivo son: {", ".join(df.columns)}')
                    return redirect('cargar_excel')

                # Procesar los datos del archivo Excel e insertarlos en la base de datos
                from .models import Capacity  # Asegúrate de importar el modelo adecuado
                
                for _, row in df.iterrows():
                    Capacity.objects.create(
                        fecha_extract_registro=row['FECHA_EXTRACT_REGISTRO'],
                        nombre_tecnico=row['NOMBRE_TECNICO'],
                        nombre_team=row['NOMBRE_TEAM'],
                        especialidad_usuario=row['ESPECIALIDAD_USUARIO'],
                        area=row['AREA'],
                        rut_tecnico=row['RUT_TECNICO'],
                        zona_operacional=row['ZONA_OPERACIONAL'],
                        zona_adjudicacion=row['ZONA_ADJUDICACION'],
                        especialidad_team=row['ESPECIALIDAD_TEAM'],
                        capacidad=row['CAPACIDAD'],
                        total_hh=row['TOTAL_HH'],
                        phi=row['PHI'],
                        total_hh_phi=row['TOTAL_HH_PHI'],
                    )

            except pd.errors.ParserError as e:
                messages.error(request, f'Error al leer el archivo Excel: {str(e)}')
                return redirect('cargar_excel')
            except Exception as e:
                messages.error(request, f'Error inesperado al procesar el archivo Excel: {str(e)}')
                return redirect('cargar_excel')

            messages.success(request, 'Archivo Excel cargado correctamente')
            return redirect('cargar_excel')
    else:
        form = ExcelUploadForm()
    return render(request, 'pagos/cargar_excel.html', {'form': form})



@login_required
def generar_tabla_pago_sla_mes(request, year, month):
    """
    Vista para generar la tabla de pago SLA por mes y año.
    """
    if request.method == 'POST':
        form = GenerarTablaSLAMesForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            logger.info(f"Tabla SLA generada para {month}/{year} por {request.user.username}.")
            messages.success(request, 'Tabla SLA generada exitosamente.')
            return redirect('sla_dashboard')  # Redirige a la vista del dashboard SLA
        else:
            logger.warning(f"Errores en el formulario SLA: {form.errors}")
            messages.error(request, 'Hubo errores en el formulario. Por favor, corrígelos.')
    else:
        form = GenerarTablaSLAMesForm(initial={'fecha': f"{year}-{month}-01"})  # Ajusta según tus necesidades
    
    context = {
        'form': form,
        'year': year,
        'month': month,
    }
    return render(request, 'pagos/generar_tabla_pago_sla_mes.html', context)



@login_required
def cargar_tabla_lpu(request):
    if request.method == 'POST':
        form = LPUExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            try:
                df = pd.read_excel(excel_file)
                df.columns = df.columns.str.strip()
                
                mes_mapping = {
                    'ene': 1, 'feb': 2, 'mar': 3, 'abr': 4, 'may': 5, 'jun': 6,
                    'jul': 7, 'ago': 8, 'sept': 9, 'oct': 10, 'nov': 11, 'dic': 12
                }

                for _, row in df.iterrows():
                    try:
                        pagada_value = row['PAGADA'].strip().lower() if not pd.isna(row['PAGADA']) else 'no'
                        pagada = pagada_value in ['yes', 'si', 'true', '1', 'sí']
                        
                        tiempo_trans_sap = row['TIEMPO TRANS. SAP'] if not pd.isna(row['TIEMPO TRANS. SAP']) else 0
                        tiempo_validar_trans_sap = row['TIEMPO VALIDAR TRANS. SAP'] if not pd.isna(row['TIEMPO VALIDAR TRANS. SAP']) else 0
                        mes_finalizacion = mes_mapping.get(str(row['mes finalización']).strip().lower(), None)
                        
                        precio_mdm = row['PRECIO MDM'] if not pd.isna(row['PRECIO MDM']) else 0
                        factor_multiplicador = row['FACTOR MULTIPLICADOR'] if not pd.isna(row['FACTOR MULTIPLICADOR']) else 1
                        total = row['TOTAL'] if not pd.isna(row['TOTAL']) else 0
                        
                        fecha_finalizacion_tarea = pd.to_datetime(row['FECHA FINALIZACIÓN TAREA'], errors='coerce')
                        fecha_ingreso_trans_sap = pd.to_datetime(row['FECHA INGRESO TRANS. SAP'], errors='coerce')
                        fecha_ingreso_validar_trans_sap = pd.to_datetime(row['FECHA INGRESO VALIDAR TRANS. SAP'], errors='coerce')
                        fecha_final = pd.to_datetime(row['FECHA FINAL'], errors='coerce')

                        fecha_finalizacion_tarea = None if pd.isna(fecha_finalizacion_tarea) else fecha_finalizacion_tarea
                        fecha_ingreso_trans_sap = None if pd.isna(fecha_ingreso_trans_sap) else fecha_ingreso_trans_sap
                        fecha_ingreso_validar_trans_sap = None if pd.isna(fecha_ingreso_validar_trans_sap) else fecha_ingreso_validar_trans_sap
                        fecha_final = None if pd.isna(fecha_final) else fecha_final

                        tipo_red2 = "MOVIL" if "MOVIL" in row['TIPO RED'].upper() else "FIJA"

                        LPU.objects.create(
                            numero_de_tarea=row['NUMERO DE TAREA'],
                            tipo_red=row['TIPO RED'],
                            tipo_red2=tipo_red2,  # Asignación de tipo_red2 basada en tipo_red
                            area_empresa=row['ÁREA/EMPRESA'],
                            zona_adjudicacion=row['ZONA ADJUDICACIÓN'],
                            zona_operativa=row['ZONA OPERATIVA'],
                            zona_cluster=row['ZONA CLUSTER'],
                            estado_de_trabajo=row['ESTADO DE TRABAJO'],
                            sap_fijo_lpu=row['SAP FIJO LPU'],
                            sap_capex_lpu=row['SAP CAPEX LPU'],
                            sap_opex_lpu=row['SAP OPEX LPU'],
                            item_lpu=row['ITEM LPU'],
                            precio_mdm=precio_mdm,
                            servicios=row['SERVICIOS'],
                            factor_multiplicador=factor_multiplicador,
                            total=total,
                            estado_presupuesto=row['ESTADO PRESUPUESTO'],
                            pagada=pagada,
                            fecha_finalizacion_tarea=fecha_finalizacion_tarea,
                            fecha_ingreso_trans_sap=fecha_ingreso_trans_sap,
                            tiempo_trans_sap=tiempo_trans_sap,
                            fecha_ingreso_validar_trans_sap=fecha_ingreso_validar_trans_sap,
                            tiempo_validar_trans_sap=tiempo_validar_trans_sap,
                            fecha_final=fecha_final,
                            nombre_de_proyecto=row['NOMBRE DE PROYECTO'],
                            ano_finalizacion=row['año finalizacion'],
                            mes_finalizacion=mes_finalizacion,
                            observaciones=row.get('OBSERVACIONES', '')
                        )

                    except Exception as e:
                        logger.error(f"Error creating LPU entry: {e}")
                        continue

                messages.success(request, 'Archivo cargado exitosamente y tipo_red2 actualizado.')
                return redirect('menu')
            except Exception as e:
                logger.error(f'Error loading Excel file: {e}')
                messages.error(request, f'Error al cargar el archivo: {e}')
        else:
            messages.error(request, 'Por favor seleccione un archivo válido.')
    else:
        form = LPUExcelUploadForm()

    return render(request, 'pagos/cargar_tabla_lpu.html', {'form': form})

# Vista para ver los registros LPU
@login_required
def ver_lpu(request):
    # Obtiene todos los registros de LPU, ordenados por el campo 'id' en orden ascendente
    lpu_list = LPU.objects.all().order_by('id')
    
    # Crea un paginador que divide los registros en páginas de 100 elementos cada una
    paginator = Paginator(lpu_list, 100)
    
    # Obtiene el número de página actual de los parámetros de la solicitud GET
    page_number = request.GET.get('page')
    
    # Obtiene el conjunto de objetos correspondientes a la página actual
    page_obj = paginator.get_page(page_number)
    
    # Renderiza la plantilla 'ver_lpu.html' pasando la página actual (page_obj) al contexto
    return render(request, 'pagos/ver_lpu.html', {'page_obj': page_obj})



@login_required
def descargar_capacity(request):
    # Obtener todos los registros de la tabla Capacity
    capacity_data = Capacity.objects.all()

    # Crear una lista de diccionarios con los datos que quieres exportar
    data = []
    for record in capacity_data:
        data.append({
            'Fecha Extract Registro': record.fecha_extract_registro,
            'Nombre Técnico': record.nombre_tecnico,
            'Nombre Team': record.nombre_team,
            'Especialidad Usuario': record.especialidad_usuario,
            'Área': record.area,
            'RUT Técnico': record.rut_tecnico,
            'Zona Operacional': record.zona_operacional,
            'Zona Adjudicación': record.zona_adjudicacion,
            'Especialidad Team': record.especialidad_team,
            'Capacidad': record.capacidad,
            'Total HH': record.total_hh,
            'PHI': record.phi,
            'Total HH + PHI': record.total_hh_phi
        })

    # Convertir la lista de diccionarios en un DataFrame de pandas
    df = pd.DataFrame(data)

    # Crear una respuesta HTTP con el tipo de contenido para archivos Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=capacity_data.xlsx'

    # Guardar el DataFrame en el archivo Excel
    df.to_excel(response, index=False)

    return response


@login_required
def obtener_precio_especialidad(especialidad, zona_adjudicacion, zona_operacional, area):
    # Obtener el año actual
    anio_actual = datetime.now().year

    # Buscar el precio para el año actual
    try:
        precio = PrecioEspecialidad.objects.get(
            especialidad=especialidad,
            zona_adjudicacion=zona_adjudicacion,
            zona_operacional=zona_operacional,
            area=area,
            anio=anio_actual
        )
        return precio.precio
    except PrecioEspecialidad.DoesNotExist:
        return Decimal('0.00')  # Retornar 0 si no hay precio disponible


@login_required
def actualizar_precios_desde_excel(archivo_excel):
    # Leer el archivo Excel con los nuevos precios
    df = pd.read_excel(archivo_excel)

    for _, row in df.iterrows():
        # Por cada fila en el Excel, actualizar o crear un nuevo precio
        especialidad = row['especialidad']
        zona_adjudicacion = row['zona_adjudicacion']
        zona_operacional = row['zona_operacional']
        area = row['area']
        precio = Decimal(row['nuevo_precio_2024'])
        anio = row['anio']

        # Crear o actualizar el precio para la especialidad
        PrecioEspecialidad.objects.update_or_create(
            especialidad=especialidad,
            zona_adjudicacion=zona_adjudicacion,
            zona_operacional=zona_operacional,
            area=area,
            anio=anio,
            defaults={'precio': precio}
        )


@login_required
def generar_pagos_para_capacity(capacity_data):
    for capacity in capacity_data:
        especialidad = capacity.especialidad_usuario
        zona_adjudicacion = capacity.zona_adjudicacion
        zona_operacional = capacity.zona_operacional
        area = capacity.area

        # Obtener el precio actual
        precio_actual = obtener_precio_especialidad(especialidad, zona_adjudicacion, zona_operacional, area)

        # Calcular el pago basado en el precio actual y otros datos del capacity
        total_a_pagar = precio_actual * capacity.Q  # Ejemplo, dependiendo de cómo se calculan los pagos

        # Guardar el total a pagar en la tabla
        capacity.total_a_pagar = total_a_pagar
        capacity.save()



# Vista para el dashboard de LPU (Lista de Precios Unitarios)
import calendar
from django.utils.translation import gettext as _
from django.db.models import Count

@login_required
def dashboard_lpu(request):
    # Mapeo de abreviaturas de meses a números
    mes_mapping = {
        'ene': 1, 'feb': 2, 'mar': 3, 'abr': 4, 'may': 5, 'jun': 6,
        'jul': 7, 'ago': 8, 'sept': 9, 'oct': 10, 'nov': 11, 'dic': 12
    }

    # Obtiene los datos de LPU agrupados por mes y año de finalización, y cuenta el número total de registros por grupo
    lpu_data = (
        LPU.objects
        .values('mes_finalizacion', 'ano_finalizacion')  # Selecciona los campos de mes y año de finalización
        .annotate(total_registros=Count('id'))  # Agrega una anotación que cuenta el total de registros por cada combinación de mes y año
        .order_by('ano_finalizacion', 'mes_finalizacion')  # Ordena los resultados por año y mes de finalización
    )

    # Traducir los números de los meses a nombres de meses en texto
    for record in lpu_data:
        mes_finalizacion = record['mes_finalizacion']
        if isinstance(mes_finalizacion, str):  # Si el mes es un texto (abreviatura)
            mes_numero = mes_mapping.get(mes_finalizacion.lower())
        else:  # Si el mes ya es un número
            mes_numero = mes_finalizacion

        if mes_numero:
            # Si mes_numero es válido, traduce el número del mes al nombre completo del mes
            record['mes_nombre'] = _(calendar.month_name[mes_numero])
        else:
            # Si no hay un mes válido, asigna None
            record['mes_nombre'] = None

    # Define el contexto a pasar a la plantilla, con los datos LPU procesados
    context = {
        'lpu_data': lpu_data,
    }

    # Renderiza la plantilla 'dashboard_lpu.html' pasando el contexto con los datos
    return render(request, 'pagos/dashboard_lpu.html', context)

import openpyxl
from django.http import HttpResponse
from .models import LPU

import openpyxl
from django.http import HttpResponse
from .models import LPU

def download_lpu_excel(request):
    # Crear un libro de trabajo en Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LPU Data"

    # Definir los encabezados de la tabla
    headers = ['NUMERO DE TAREA', 'TIPO RED', 'ÁREA/EMPRESA', 'ZONA ADJUDICACIÓN', 'ZONA OPERATIVA',
               'ZONA CLUSTER', 'ESTADO DE TRABAJO', 'SAP FIJO LPU', 'SAP CAPEX LPU', 'SAP OPEX LPU',
               'ITEM LPU', 'PRECIO MDM', 'SERVICIOS', 'FACTOR MULTIPLICADOR', 'TOTAL', 'ESTADO PRESUPUESTO',
               'PAGADA', 'FECHA FINALIZACIÓN TAREA', 'FECHA INGRESO TRANS. SAP', 'TIEMPO TRANS. SAP',
               'FECHA INGRESO VALIDAR TRANS. SAP', 'TIEMPO VALIDAR TRANS. SAP', 'FECHA FINAL',
               'NOMBRE DE PROYECTO', 'AÑO FINALIZACION', 'MES FINALIZACION', 'TIPO RED2', 'OBSERVACIONES']
    ws.append(headers)

    # Obtener los datos de LPU y añadir cada fila a la hoja
    for lpu in LPU.objects.all():
        row = [
            lpu.numero_de_tarea, lpu.tipo_red, lpu.area_empresa, lpu.zona_adjudicacion,
            lpu.zona_operativa, lpu.zona_cluster, lpu.estado_de_trabajo, lpu.sap_fijo_lpu,
            lpu.sap_capex_lpu, lpu.sap_opex_lpu, lpu.item_lpu, 
            round(lpu.precio_mdm, 2) if lpu.precio_mdm is not None else None,  # Formatear a dos decimales
            lpu.servicios, lpu.factor_multiplicador, lpu.total, lpu.estado_presupuesto,
            'SI' if lpu.pagada else 'NO', lpu.fecha_finalizacion_tarea, lpu.fecha_ingreso_trans_sap, 
            lpu.tiempo_trans_sap, lpu.fecha_ingreso_validar_trans_sap, lpu.tiempo_validar_trans_sap, 
            lpu.fecha_final, lpu.nombre_de_proyecto, lpu.ano_finalizacion, 
            lpu.mes_finalizacion, lpu.tipo_red2, lpu.observaciones
        ]
        ws.append(row)

    # Aplicar formato de número a la columna "PRECIO MDM" (columna 12 en Excel)
    for cell in ws['L'][1:]:  # Saltamos el encabezado
        cell.number_format = '#,##0.00'  # Formato de dos decimales

    # Configurar la respuesta HTTP para la descarga de Excel
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename="lpu_data.xlsx"'
    wb.save(response)

    return response



# Vista para ver registros LPU filtrados por mes y año específicos
@login_required
def ver_lpu_por_mes(request, year, month):
    # Filtra los registros de LPU por el año y mes de finalización especificados en los parámetros
    lpu_data = LPU.objects.filter(ano_finalizacion=year, mes_finalizacion=month)
    
    # Crea un paginador que divide los resultados en páginas de 10 registros
    paginator = Paginator(lpu_data, 10)  # Por ejemplo, 10 registros por página
    
    # Obtiene el número de página actual de los parámetros de la solicitud GET
    page_number = request.GET.get('page')
    
    # Obtiene el conjunto de objetos correspondientes a la página actual
    page_obj = paginator.get_page(page_number)

    # Define el contexto que se pasará a la plantilla, incluyendo el objeto de la página actual y el mes/año de finalización
    context = {
        'page_obj': page_obj,
        'mes_finalizacion': f"{year}-{month}"  # Combina año y mes para mostrarlos juntos en la plantilla
    }
    
    # Renderiza la plantilla 'lpu_por_mes.html' pasando el contexto con los datos filtrados
    return render(request, 'pagos/lpu_por_mes.html', context)




# Vista para crear un nuevo registro LPU a través de un formulario HTML
@login_required
def crear_lpu(request):
    # Si el método de la solicitud es POST (el formulario ha sido enviado)
    if request.method == 'POST':
        form = LPUForm(request.POST)
        if form.is_valid():
            # Guarda el nuevo registro LPU en la base de datos
            form.save()
            # Mensaje de éxito
            messages.success(request, 'Registro LPU creado exitosamente.')
            return redirect('ver_lpu')
        else:
            # Mensaje de error si el formulario no es válido
            messages.error(request, 'Por favor corrija los errores en el formulario.')
    else:
        form = LPUForm()

    # Obtener valores únicos para los campos desplegables
    numero_de_tarea_choices = LPU.objects.values_list('numero_de_tarea', flat=True).distinct()
    tipo_red_choices = LPU.objects.values_list('tipo_red', flat=True).distinct()
    area_empresa_choices = LPU.objects.values_list('area_empresa', flat=True).distinct()
    zona_adjudicacion_choices = LPU.objects.values_list('zona_adjudicacion', flat=True).distinct()
    zona_operativa_choices = LPU.objects.values_list('zona_operativa', flat=True).distinct()
    nombre_de_proyecto_choices = LPU.objects.values_list('nombre_de_proyecto', flat=True).distinct()

    # Contexto para la plantilla
    context = {
        'form': form,
        'numero_de_tarea_choices': numero_de_tarea_choices,
        'tipo_red_choices': tipo_red_choices,
        'area_empresa_choices': area_empresa_choices,
        'zona_adjudicacion_choices': zona_adjudicacion_choices,
        'zona_operativa_choices': zona_operativa_choices,
        'nombre_de_proyecto_choices': nombre_de_proyecto_choices,
    }

    # Renderiza la plantilla 'crear_lpu.html' con el formulario y los valores únicos
    return render(request, 'pagos/crear_lpu.html', context)


# Vista para editar un registro LPU en la base de datos desde una plantilla HTML
@login_required
def editar_lpu(request, pk):
    # Obtiene el objeto LPU correspondiente al id (pk) o devuelve un error 404 si no existe
    lpu = get_object_or_404(LPU, pk=pk)
    
    # Si el método de la solicitud es POST (el formulario ha sido enviado)
    if request.method == 'POST':
        # Crea una instancia del formulario con los datos enviados y asocia el registro existente (instance=lpu)
        form = LPUForm(request.POST, instance=lpu)
        # Verifica si el formulario es válido
        if form.is_valid():
            # Si es válido, guarda los cambios en el registro LPU en la base de datos
            form.save()
            # Muestra un mensaje de éxito al usuario
            messages.success(request, 'El registro de LPU se ha actualizado correctamente.')
            # Redirige a la vista 'ver_lpu' después de guardar los cambios
            return redirect('ver_lpu')
        else:
            # Si el formulario no es válido, muestra un mensaje de error
            messages.error(request, 'Hubo un error al actualizar el registro. Por favor revisa el formulario.')
            # Imprime los errores del formulario en la consola para depuración
            print(form.errors)
    else:
        # Si el método no es POST (es decir, es GET), muestra el formulario con los datos del registro LPU existente
        form = LPUForm(instance=lpu)
    
    # Renderiza la plantilla 'editar_lpu.html' pasando el formulario con los datos del registro LPU
    return render(request, 'pagos/editar_lpu.html', {'form': form})



# Vista para eliminar un registro LPU de la base de datos
@login_required
def eliminar_lpu(request, pk):
    # Obtiene el objeto LPU correspondiente al id (pk) o devuelve un error 404 si no existe
    lpu = get_object_or_404(LPU, pk=pk)
    
    # Verifica si la solicitud es de tipo POST (es decir, el usuario ha confirmado la eliminación)
    if request.method == 'POST':
        # Elimina el registro de LPU de la base de datos
        lpu.delete()
        # Redirige a la vista 'ver_lpu' (la lista de registros) después de eliminar el registro
        return redirect('ver_lpu')
    
    # Si el método no es POST (es decir, es GET), se renderiza una página de confirmación
    return render(request, 'pagos/eliminar_lpu.html', {'lpu': lpu})  # Muestra la página de confirmación de eliminación


class CapacityViewSet(viewsets.ModelViewSet):
    queryset = Capacity.objects.all()
    serializer_class = CapacitySerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['nombre_tecnico', 'rut_tecnico']



class NuevaTablaPagosViewSet(viewsets.ModelViewSet):
    queryset = NuevaTablaPagos.objects.all()
    serializer_class = NuevaTablaPagosSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['especialidad', 'zona_adjudicacion']
    

class TablaDePagoCapacityViewSet(viewsets.ModelViewSet):
    queryset = TablaDePagoCapacity.objects.all()
    serializer_class = TablaDePagoCapacitySerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['nombre_tecnico', 'rut_tecnico']
    


class SLAReportViewSet(viewsets.ModelViewSet):
    queryset = SLAReport.objects.all()
    serializer_class = SLAReportSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['zona_operacional', 'proyecto']
    


# API View para listar los datos de Capacity con autenticación

class CapacityListView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        capacities = Capacity.objects.all()
        paginator = Paginator(capacities, 10)  # Paginación
        page_number = request.query_params.get('page', 1)
        page_obj = paginator.get_page(page_number)
        serializer = CapacitySerializer(page_obj, many=True)
        return Response(serializer.data)




# API View para obtener los detalles de un registro de Capacity por su ID

class CapacityDetailView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        capacity = get_object_or_404(Capacity, pk=pk)
        serializer = CapacitySerializer(capacity)
        return Response(serializer.data)




# API View para listar los datos de NuevaTablaPagos con autenticación

class NuevaTablaPagosListView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        pagos = NuevaTablaPagos.objects.all()
        paginator = Paginator(pagos, 10)
        page_number = request.query_params.get('page', 1)
        page_obj = paginator.get_page(page_number)
        serializer = NuevaTablaPagosSerializer(page_obj, many=True)
        return Response(serializer.data)




# API View para listar los datos de TablaDePagoCapacity con autenticación

class TablaDePagoCapacityListView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        pagos_capacity = TablaDePagoCapacity.objects.all()
        paginator = Paginator(pagos_capacity, 10)
        page_number = request.query_params.get('page', 1)
        page_obj = paginator.get_page(page_number)
        serializer = TablaDePagoCapacitySerializer(page_obj, many=True)
        return Response(serializer.data)




# API View para listar los datos de SLAReport con autenticación

class SLAReportListView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        sla_reports = SLAReport.objects.all()
        paginator = Paginator(sla_reports, 10)
        page_number = request.query_params.get('page', 1)
        page_obj = paginator.get_page(page_number)
        serializer = SLAReportSerializer(page_obj, many=True)
        return Response(serializer.data)




# Vista para ver los registros de la tabla NuevaTablaPagos
@login_required
def ver_nuevatabladepagos(request):
    # Obtiene todos los registros de NuevaTablaPagos de la base de datos
    datos_list = NuevaTablaPagos.objects.all()
    
    # Imprime en la consola el número de registros obtenidos, útil para depuración
    print(f"Datos obtenidos: {datos_list.count()} registros")  # Verificar que hay datos

    # Paginación: divide los registros en páginas de 10 elementos cada una
    paginator = Paginator(datos_list, 10)  # 10 registros por página
    # Obtiene el número de página de los parámetros GET
    page = request.GET.get('page')
    
    try:
        # Intenta obtener los datos de la página solicitada
        datos = paginator.page(page)
    except PageNotAnInteger:
        # Si el número de página no es un entero, carga la primera página
        datos = paginator.page(1)
    except EmptyPage:
        # Si el número de página es mayor que el número de páginas, carga la última página
        datos = paginator.page(paginator.num_pages)

    # Define el contexto a pasar a la plantilla, con los datos paginados
    context = {
        'datos': datos  # Asegúrate de que 'datos' sea la clave que estás usando en la plantilla
    }

    # Renderiza la plantilla 'ver_nueva_tabla_pagos.html' con los datos paginados
    return render(request, 'pagos/ver_nueva_tabla_pagos.html', context)



# Vista para buscar registros en la tabla NuevaTablaPagos
@login_required
def buscar_nueva_tabla_pagos(request):
    # Obtiene el valor de búsqueda de los parámetros GET (q es el parámetro de búsqueda)
    query = request.GET.get('q', '')  # Si no se pasa ningún valor, el valor predeterminado es una cadena vacía

    if query:
        # Si hay una búsqueda, filtra los registros de NuevaTablaPagos en función de los campos especificados
        datos = NuevaTablaPagos.objects.filter(
            sociedad__icontains=query  # Filtra por coincidencias parciales en el campo 'sociedad'
        ) | NuevaTablaPagos.objects.filter(
            material__icontains=query  # Filtra por coincidencias parciales en el campo 'material'
        ) | NuevaTablaPagos.objects.filter(
            especialidad__icontains=query  # Filtra por coincidencias parciales en el campo 'especialidad'
        ) | NuevaTablaPagos.objects.filter(
            zona_adjudicacion__icontains=query  # Filtra por coincidencias parciales en el campo 'zona_adjudicacion'
        ) | NuevaTablaPagos.objects.filter(
            zona_operacional__icontains=query  # Filtra por coincidencias parciales en el campo 'zona_operacional'
        )
    else:
        # Si no hay término de búsqueda, recupera todos los registros de la tabla
        datos = NuevaTablaPagos.objects.all()

    # Renderiza la plantilla 'ver_nueva_tabla_pagos.html', pasando los resultados filtrados y el término de búsqueda
    return render(request, 'pagos/ver_nueva_tabla_pagos.html', {'datos': datos, 'query': query})



# Vista para editar un registro en la tabla NuevaTablaPagos
@login_required
def editar_nuevatabladepagos(request, pk):
    # Obtiene el registro de NuevaTablaPagos correspondiente al ID `pk` o devuelve un error 404 si no existe
    registro = get_object_or_404(NuevaTablaPagos, pk=pk)

    if request.method == 'POST':
        # Crea una instancia del formulario con los datos enviados y asocia el registro existente
        form = NuevaTablaPagosForm(request.POST, instance=registro)
        if form.is_valid():
            # Guarda los cambios en el registro existente
            form.save()
            # Muestra un mensaje de éxito al usuario
            messages.success(request, 'Registro actualizado exitosamente.')
            # Redirige al usuario a la vista 'nuevatabladepagos'
            return redirect('nuevatabladepagos')
        else:
            # Muestra un mensaje de error si hay errores en el formulario
            messages.error(request, 'Por favor, corrija los errores en el formulario.')
    else:
        # Si el método no es POST, muestra el formulario con los datos actuales del registro
        form = NuevaTablaPagosForm(instance=registro)

    # Listas desplegables dinámicas basadas en valores únicos de la base de datos
    sociedad_choices = NuevaTablaPagos.objects.values_list('sociedad', flat=True).distinct()
    material_choices = NuevaTablaPagos.objects.values_list('material', flat=True).distinct()
    especialidad_choices = NuevaTablaPagos.objects.values_list('especialidad', flat=True).distinct()
    zona_adjudicacion_choices = NuevaTablaPagos.objects.values_list('zona_adjudicacion', flat=True).distinct()
    zona_operacional_choices = NuevaTablaPagos.objects.values_list('zona_operacional', flat=True).distinct()
    area_choices = NuevaTablaPagos.objects.values_list('area', flat=True).distinct()

    # Contexto para la plantilla
    context = {
        'form': form,
        'registro': registro,
        'sociedad_choices': sociedad_choices,
        'material_choices': material_choices,
        'especialidad_choices': especialidad_choices,
        'zona_adjudicacion_choices': zona_adjudicacion_choices,
        'zona_operacional_choices': zona_operacional_choices,
        'area_choices': area_choices,
    }

    # Renderiza la plantilla con el formulario y las listas desplegables dinámicas
    return render(request, 'pagos/editar_nuevatabladepagos.html', context)




# Vista para eliminar un registro de la tabla NuevaTablaPagos
@login_required
def eliminar_nuevatabladepagos(request, pk):
    # Obtiene el registro de NuevaTablaPagos correspondiente al ID `pk` o devuelve un error 404 si no existe
    registro = get_object_or_404(NuevaTablaPagos, pk=pk)
    
    # Si el método de la solicitud es POST (es decir, el usuario ha confirmado la eliminación)
    if request.method == 'POST':
        # Elimina el registro de la base de datos
        registro.delete()
        # Redirige a la lista de registros 'nuevatabladepagos' después de eliminar el registro
        return redirect('nuevatabladepagos')
    
    # Si el método no es POST (es decir, es GET), muestra una página de confirmación antes de eliminar
    return render(request, 'pagos/eliminar_nuevatabladepagos.html', {'registro': registro})


@login_required
def crear_nuevatabladepagos(request):
    if request.method == 'POST':
        # Crea una instancia del formulario con los datos enviados por el usuario
        form = NuevaTablaPagosForm(request.POST)
        if form.is_valid():
            # Guarda el nuevo registro en la base de datos
            form.save()
            # Muestra un mensaje de éxito al usuario
            messages.success(request, 'Registro creado exitosamente.')
            # Redirige a la vista 'nuevatabladepagos'
            return redirect('nuevatabladepagos')
        else:
            # Muestra un mensaje de error si el formulario no es válido
            messages.error(request, 'Por favor, corrija los errores en el formulario.')
    else:
        # Si el método no es POST, inicializa un formulario vacío
        form = NuevaTablaPagosForm()

    # Listas desplegables dinámicas basadas en valores únicos de la base de datos
    sociedad_choices = NuevaTablaPagos.objects.values_list('sociedad', flat=True).distinct()
    material_choices = NuevaTablaPagos.objects.values_list('material', flat=True).distinct()
    especialidad_choices = NuevaTablaPagos.objects.values_list('especialidad', flat=True).distinct()
    zona_adjudicacion_choices = NuevaTablaPagos.objects.values_list('zona_adjudicacion', flat=True).distinct()
    zona_operacional_choices = NuevaTablaPagos.objects.values_list('zona_operacional', flat=True).distinct()
    area_choices = NuevaTablaPagos.objects.values_list('area', flat=True).distinct()

    # Contexto para la plantilla
    context = {
        'form': form,
        'sociedad_choices': sociedad_choices,
        'material_choices': material_choices,
        'especialidad_choices': especialidad_choices,
        'zona_adjudicacion_choices': zona_adjudicacion_choices,
        'zona_operacional_choices': zona_operacional_choices,
        'area_choices': area_choices,
    }

    # Renderiza la plantilla con el formulario y las listas desplegables dinámicas
    return render(request, 'pagos/crear_nueva_tabla_pagos.html', context)

from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from .models import TablaDePagoCapacity
from datetime import datetime
from django.db.models import Sum, F, Count
import pandas as pd

# Vista para el dashboard
from django.db.models import F, Count

from django.db.models import F, Count
from django.shortcuts import render

def dashboard_tabladepago_capacity(request):
    # Agrupa registros por mes y año utilizando el nuevo campo `fecha`
    records = (
        TablaDePagoCapacity.objects
        .annotate(month=F('fecha__month'), year=F('fecha__year'))
        .values('month', 'year')
        .annotate(total_records=Count('id'))
        .order_by('year', 'month')
    )
    context = {
        'records': records,
    }
    return render(request, 'pagos/dashboard_tabladepago_capacity.html', context)

def ver_tabla_pago_capacity_mes(request, year, month):
    # Filtra los registros por el mes y el año de la fecha
    registros = TablaDePagoCapacity.objects.filter(
        fecha__year=year,
        fecha__month=month
    )
    context = {
        'registros': registros,
        'year': year,
        'month': month,
    }
    return render(request, 'pagos/pago_capacity_mes.html', context)

# Vista para descargar registros por mes como Excel
def descargar_tabla_pago_capacity_mes(request, year, month):
    registros = TablaDePagoCapacity.objects.filter(
        fecha__year=year, 
        fecha__month=month
    ).values()
    df = pd.DataFrame(list(registros))

    # Crear el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=TablaPagoCapacity_{year}_{month}.xlsx'
    df.to_excel(response, index=False)
    return response


from django.contrib.auth.decorators import login_required
from django.db.models import Count, Sum
from django.db.models.functions import TruncMonth
from django.shortcuts import render
from .models import Capacity

@login_required
def dashboard_capacity(request):
    # Agrupar los registros por mes en función de `fecha_extract_registro`
    records = (
        Capacity.objects
        .annotate(month=TruncMonth('fecha_extract_registro'))
        .values('month')
        .annotate(total_records=Count('id'))
        .order_by('month')
    )

    # Agrupar los totales a pagar por zona, especialidad y área, sumando `total_hh_phi`
    totals = (
        Capacity.objects
        .values('zona_operacional', 'especialidad_usuario', 'area')
        .annotate(total_a_pago=Sum('total_hh_phi'))
    )

    # Pasar los datos al contexto de la plantilla
    context = {
        'records': records,
        'totals': totals,
    }

    return render(request, 'pagos/dashboard_capacity.html', context)



# Vista para ver todos los datos del modelo Capacity con opción de búsqueda y paginación
@login_required
def ver_todos_los_datos_capacity(request):
    # Obtiene el valor del campo de búsqueda desde los parámetros GET (q es el término de búsqueda)
    query = request.GET.get('q', '')  # Si no se pasa un valor, usa una cadena vacía como valor predeterminado

    # Filtra los datos en función de los términos de búsqueda, si hay una búsqueda
    if query:
        # Busca coincidencias parciales en varios campos del modelo Capacity
        datos_list = Capacity.objects.filter(
            nombre_tecnico__icontains=query  # Coincidencias parciales en el campo 'nombre_tecnico'
        ) | Capacity.objects.filter(
            especialidad_usuario__icontains=query  # Coincidencias parciales en 'especialidad_usuario'
        ) | Capacity.objects.filter(
            area__icontains=query  # Coincidencias parciales en 'area'
        ) | Capacity.objects.filter(
            zona_operacional__icontains=query  # Coincidencias parciales en 'zona_operacional'
        )
    else:
        # Si no hay término de búsqueda, recupera todos los registros del modelo Capacity
        datos_list = Capacity.objects.all()

    # Ordena los registros por el campo 'fecha_extract_registro'
    datos_list = datos_list.order_by('fecha_extract_registro')
    
    # Paginación: divide los resultados en páginas de 10 registros
    paginator = Paginator(datos_list, 10)  # 10 registros por página
    # Obtiene el número de página actual de los parámetros GET
    page_number = request.GET.get('page')
    # Obtiene los registros correspondientes a la página solicitada
    datos_paginados = paginator.get_page(page_number)

    # Renderiza la plantilla 'ver_datos_capacity_todos.html', pasando los datos paginados y la búsqueda actual
    return render(request, 'pagos/ver_datos_capacity_todos.html', {
        'datos': datos_paginados,  # Los registros paginados
        'query': query,  # El término de búsqueda actual, para mantener el valor en el campo de búsqueda
    })



@login_required
def descargar_capacity_por_mes(request, year, month):
    # Filtra los registros del modelo Capacity por el año y mes especificados
    datos = Capacity.objects.filter(
        fecha_extract_registro__year=year,
        fecha_extract_registro__month=month
    )
    
    # Si no hay datos para el mes y año especificados, devolver un mensaje de error
    if not datos.exists():
        messages.error(request, f"No se encontraron registros para {month}/{year}.")
        return redirect('dashboard_capacity')
    
    # Crear un DataFrame con los datos filtrados
    df = pd.DataFrame(list(datos.values(
        'fecha_extract_registro', 'nombre_tecnico', 'nombre_team', 'especialidad_usuario',
        'area', 'rut_tecnico', 'zona_operacional', 'zona_adjudicacion',
        'especialidad_team', 'capacidad', 'total_hh', 'phi', 'total_hh_phi'
    )))
    
    # Crear una respuesta HTTP con el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Capacity_{year}_{month}.xlsx'
    
    # Guardar el DataFrame en un archivo Excel
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=f'Capacity_{year}_{month}')
    
    # Retornar la respuesta con el archivo Excel descargable
    return response



@login_required
def dashboard_precios_especialidades(request):
    # Obtener los datos agrupados por año y especialidad
    precios_por_anio = (
        NuevaTablaPagos.objects.values('anio', 'especialidad')
        .annotate(
            precio_promedio=Avg('nuevo_precio'),
            total_registros=Count('id')
        )
        .order_by('anio', 'especialidad')
    )

    # Organizar los datos en un diccionario para agrupar por año
    datos_agrupados = {}
    for item in precios_por_anio:
        anio = item['anio']
        if anio not in datos_agrupados:
            datos_agrupados[anio] = []
        datos_agrupados[anio].append({
            'especialidad': item['especialidad'],
            'precio_promedio': item['precio_promedio'],
            'total_registros': item['total_registros'],
        })

    return render(request, 'pagos/dashboard_precios_especialidades.html', {
        'datos_agrupados': datos_agrupados,
    })


@login_required
def descargar_sla(request):
    # Crear un libro de trabajo y una hoja
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'SLA Report'

    # Escribir la cabecera del Excel
    headers = ['Zona Adjudicación', 'Zona Operativa', 'Bonus o Malus', 'Proyecto', 'Fecha']
    worksheet.append(headers)

    # Obtener los datos que deseas exportar
    datos = SLAReport.objects.all()  # Obtiene todos los registros de SLAReport

    # Escribir los datos en el Excel
    for dato in datos:
        worksheet.append([
            dato.zona_adjudicacion,
            dato.zona_operativa,
            dato.bonus_o_malus,
            dato.proyecto,
            dato.fecha.strftime('%Y-%m-%d')  # Formatear la fecha
        ])

    # Guardar el archivo en un objeto BytesIO
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    # Crear la respuesta HTTP
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sla_report.xlsx"'

    return response



@login_required
def ver_detalle_especialidad(request, slug):
    # Encontrar el nombre de la especialidad a partir del slug
    especialidad_nombre = None
    especialidades = NuevaTablaPagos.objects.values_list('especialidad', flat=True).distinct()
    for esp in especialidades:
        if slugify(esp) == slug:
            especialidad_nombre = esp
            break
    
    if not especialidad_nombre:
        raise Http404("Especialidad no encontrada")
    
    registros = NuevaTablaPagos.objects.filter(especialidad=especialidad_nombre)
    return render(request, 'pagos/detalle_especialidad.html', {
        'especialidad': especialidad_nombre,
        'registros': registros,
    })



    
@login_required
def descargar_especialidad(request, especialidad):
    # Lógica para generar y descargar el archivo relacionado con la especialidad
    # Esto puede ser un archivo Excel o CSV
    # Por simplicidad, aquí solo redireccionaremos al detalle
    return redirect('ver_detalle_especialidad', especialidad=especialidad)



@login_required
def ver_todos_precios_especialidades(request):
    # Aquí puedes definir la lógica para obtener y mostrar todos los precios de especialidades
    especialidades = NuevaTablaPagos.objects.filter(anio=2024)
    return render(request, 'pagos/todos_precios_especialidades.html', {
        'especialidades': especialidades
    })



@login_required
def dashboard_precios_anuales(request):
    # Agrupar los registros de NuevaTablaPagos por año y contar el total de registros por año
    precios_anuales = NuevaTablaPagos.objects.values('anio').annotate(total_registros=Count('id')).order_by('anio')

    return render(request, 'pagos/dashboard_precios_anuales.html', {
        'precios_anuales': precios_anuales,
    })



@login_required
def sla_dashboard(request):
    """
    Vista para el dashboard de SLAReport, agrupando por mes y año.
    """
    # Agrupar los registros por mes y año
    records = SLAReport.objects.annotate(
        month=TruncMonth('fecha')
    ).values('month').annotate(
        total_records=Count('id'),
        total_bonus_malus=Sum('bonus_o_malus')
    ).order_by('-month')

    context = {
        'records': records
    }
    return render(request, 'pagos/sla_dashboard.html', context)


@login_required
def ver_todos_los_datos_sla(request):
    registros = SLAReport.objects.all()
    context = {
        'registros': registros
    }
    return render(request, 'pagos/ver_todos_los_datos_sla.html', context)


@login_required
def cargar_sla_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES['excel_file']
        
        # Verificar que el archivo sea un archivo Excel
        if not excel_file.name.endswith('.xlsx'):
            messages.error(request, 'El archivo debe ser un archivo Excel (.xlsx)')
            print("Error: El archivo no tiene la extensión .xlsx")
            return redirect('cargar_sla_excel')

        try:
            # Cargar el archivo Excel
            workbook = openpyxl.load_workbook(excel_file)
            worksheet = workbook.active
            print("Archivo Excel cargado correctamente.")
        except Exception as e:
            messages.error(request, "Error al abrir el archivo Excel.")
            print(f"Error al abrir el archivo Excel: {e}")
            return redirect('cargar_sla_excel')

        # Leer los datos del archivo Excel y guardarlos en la base de datos
        filas_procesadas = 0
        filas_omitidas = 0
        filas_exitosas = 0
        for row in worksheet.iter_rows(min_row=2, values_only=True):  # Comienza desde la segunda fila
            zona_adjudicacion, zona_operativa, bonus_o_malus, proyecto, fecha = row
            
            # Depuración: Mostrar el contenido de la fila actual
            print(f"Procesando fila: {row}")

            # Verificar si los datos están completos antes de insertar
            if None in (zona_adjudicacion, zona_operativa, bonus_o_malus, proyecto, fecha):
                filas_omitidas += 1
                print(f"Fila incompleta, omitiendo: {row}")
                continue
            
            # Convertir la fecha si está en formato de texto
            if isinstance(fecha, str):
                try:
                    fecha = datetime.strptime(fecha, '%Y-%m-%d').date()
                    print(f"Fecha convertida exitosamente: {fecha}")
                except ValueError:
                    filas_omitidas += 1
                    print(f"Fecha inválida en la fila, omitiendo: {row}")
                    continue

            # Intentar crear el registro en la base de datos
            try:
                SLAReport.objects.create(
                    zona_adjudicacion=zona_adjudicacion,
                    zona_operativa=zona_operativa,
                    bonus_o_malus=bonus_o_malus,
                    proyecto=proyecto,
                    fecha=fecha
                )
                filas_exitosas += 1
                print(f"Registro creado exitosamente para: {row}")
            except Exception as e:
                filas_omitidas += 1
                print(f"Error al crear el registro para la fila {row}: {e}")

            filas_procesadas += 1

        # Mensajes de éxito o de omisión para el usuario
        messages.success(request, f"Datos cargados exitosamente: {filas_exitosas} registros agregados.")
        messages.info(request, f"Filas procesadas: {filas_procesadas}, filas omitidas: {filas_omitidas}")

        # Resumen en consola
        print(f"Proceso completado. Filas procesadas: {filas_procesadas}, filas exitosas: {filas_exitosas}, filas omitidas: {filas_omitidas}")

        return redirect('ver_sla_report')  # Redirigir a la vista del reporte SLA

    return render(request, 'pagos/cargar_sla_excel.html')


@login_required
def descargar_sla_por_mes(request, year, month):
    registros = SLAReport.objects.filter(fecha__year=year, fecha__month=month)

    # Crear un libro de trabajo y una hoja
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = f'SLA {year}-{month}'

    # Escribir los encabezados
    headers = ['Zona Adjudicación', 'Zona Operativa', 'Bonus/Malus', 'Proyecto', 'Fecha']
    worksheet.append(headers)

    # Escribir los datos
    for registro in registros:
        worksheet.append([
            registro.zona_adjudicacion,
            registro.zona_operativa,
            registro.bonus_o_malus,
            registro.proyecto,
            registro.fecha
        ])

    # Guardar el archivo en un objeto BytesIO
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    # Crear la respuesta HTTP
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="sla_{year}_{month}.xlsx"'

    return response


@login_required
def ver_precios_por_anio(request, anio):
    """
    Vista para mostrar los precios por año.
    """
    # Obtener todos los registros de NuevaTablaPagos del año especificado
    registros = NuevaTablaPagos.objects.filter(anio=anio)

    if not registros.exists():
        # Si no hay registros para el año, mostrar un mensaje
        return render(request, 'pagos/ver_precios_por_anio.html', {
            'anio': anio,
            'mensaje': f'No se encontraron registros para el año {anio}.'
        })

    # Opcional: Calcular el precio promedio por especialidad
    precios_promedio = registros.values('especialidad').annotate(precio_promedio=Avg('nuevo_precio_2024'))

    context = {
        'anio': anio,
        'registros': registros,
        'precios_promedio': precios_promedio,
    }
    return render(request, 'pagos/ver_precios_por_anio.html', context)

@login_required
def ver_datos_sla_mes(request, year, month):
    registros = SLAReport.objects.filter(fecha__year=year, fecha__month=month)

    # Obtener el nombre del mes en español
    months = {
        1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
        5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
        9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
    }
    month_name = months.get(month, 'Mes desconocido')

    context = {
        'registros': registros,
        'year': year,
        'month_name': month_name,
    }
    return render(request, 'pagos/ver_datos_sla_mes.html', context)

@login_required
def ver_precios_por_anio(request, anio):
    # Obtener todos los registros de NuevaTablaPagos del año especificado
    registros = NuevaTablaPagos.objects.filter(anio=anio)

    if not registros.exists():
        return render(request, 'pagos/ver_precios_por_anio.html', {
            'anio': anio,
            'mensaje': f'No se encontraron registros para el año {anio}.'
        })

    return render(request, 'pagos/ver_precios_por_anio.html', {
        'anio': anio,
        'registros': registros,
    })

@login_required
def ver_datos_capacity_mes(request, year, month):
    # Filtrar por año y mes en la fecha de registro
    datos_list = Capacity.objects.filter(
        fecha_extract_registro__year=year,
        fecha_extract_registro__month=month
    ).order_by('fecha_extract_registro')

    # Capturar los parámetros de los filtros desde la URL
    especialidad_usuario = request.GET.get('especialidad_usuario', '')
    zona_operacional = request.GET.get('zona_operacional', '')
    zona_adjudicacion = request.GET.get('zona_adjudicacion', '')
    area = request.GET.get('area', '')

    # Obtener las listas únicas para los selectores de filtros
    especialidades = Capacity.objects.values_list('especialidad_usuario', flat=True).distinct()
    zonas_operacionales = Capacity.objects.values_list('zona_operacional', flat=True).distinct()
    zonas_adjudicacion = Capacity.objects.values_list('zona_adjudicacion', flat=True).distinct()
    areas = Capacity.objects.values_list('area', flat=True).distinct()

    # Aplicar los filtros si se han ingresado valores
    if especialidad_usuario:
        datos_list = datos_list.filter(especialidad_usuario__icontains=especialidad_usuario)
    if zona_operacional:
        datos_list = datos_list.filter(zona_operacional__icontains=zona_operacional)
    if zona_adjudicacion:
        datos_list = datos_list.filter(zona_adjudicacion__icontains=zona_adjudicacion)
    if area:
        datos_list = datos_list.filter(area__icontains=area)

    # Paginación
    paginator = Paginator(datos_list, 10)  # Mostrar 10 registros por página
    page_number = request.GET.get('page')
    datos = paginator.get_page(page_number)

    # Renderizar la plantilla con los datos y los filtros aplicados
    return render(request, 'pagos/ver_datos_capacity_mes.html', {
        'datos': datos,
        'especialidad_usuario': especialidad_usuario,
        'zona_operacional': zona_operacional,
        'zona_adjudicacion': zona_adjudicacion,
        'area': area,
        'especialidades': especialidades,
        'zonas_operacionales': zonas_operacionales,
        'zonas_adjudicacion': zonas_adjudicacion,
        'areas': areas,
        'year': year,
        'month': month,
    })

from django.shortcuts import render
from .models import Capacity
from .filters import CapacityFilter

@login_required
def listar_capacity(request):
    # Aplica los filtros
    filter = CapacityFilter(request.GET, queryset=Capacity.objects.all())
    # Obtiene los datos filtrados
    datos_filtrados = filter.qs

    # Paginar los datos si es necesario
    paginator = Paginator(datos_filtrados, 10)  # 10 registros por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Renderiza la plantilla con los datos filtrados y el filtro
    return render(request, 'pagos/listar_capacity.html', {'filter': filter, 'page_obj': page_obj})



# Vista para editar un registro de Capacity
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
import logging
from .models import Capacity
from .forms import CapacityForm

logger = logging.getLogger(__name__)

@login_required
def editar_capacity(request, id):
    # Obtiene el registro de Capacity a editar o lanza un error 404 si no se encuentra
    record = get_object_or_404(Capacity, pk=id)
    logger.info(f"Iniciando edición del registro con ID: {id}")

    # Si el método de la solicitud es POST (el formulario ha sido enviado)
    if request.method == 'POST':
        logger.info("Petición POST recibida, procesando formulario")
        # Crea una instancia del formulario con los datos enviados por el usuario, vinculada al registro existente
        form = CapacityForm(request.POST, instance=record)

        # Verifica si el formulario es válido (pasa todas las validaciones)
        if form.is_valid():
            form.save()
            logger.info(f"Cambios guardados exitosamente para el registro con ID: {id}")
            messages.success(request, 'Cambios guardados exitosamente.')
            return redirect('ver_datos_capacity_mes', record.fecha_extract_registro.year, record.fecha_extract_registro.month)
        else:
            logger.warning(f"Errores en el formulario: {form.errors}")
            messages.error(request, 'Por favor, corrija los errores en el formulario.')
    else:
        logger.info(f"Mostrando formulario para el registro con ID: {id}")
        form = CapacityForm(instance=record)

    # Obtener valores únicos para los campos desplegables
    nombre_team_choices = Capacity.objects.values_list('nombre_team', flat=True).distinct()
    especialidad_usuario_choices = Capacity.objects.values_list('especialidad_usuario', flat=True).distinct()
    area_choices = Capacity.objects.values_list('area', flat=True).distinct()
    zona_operacional_choices = Capacity.objects.values_list('zona_operacional', flat=True).distinct()
    especialidad_team_choices = Capacity.objects.values_list('especialidad_team', flat=True).distinct()
    zona_adjudicacion_choices = Capacity.objects.values_list('zona_adjudicacion', flat=True).distinct()

    # Contexto para renderizar la plantilla
    context = {
        'form': form,
        'record': record,
        'nombre_team_choices': nombre_team_choices,
        'especialidad_usuario_choices': especialidad_usuario_choices,
        'area_choices': area_choices,
        'zona_operacional_choices': zona_operacional_choices,
        'especialidad_team_choices': especialidad_team_choices,
        'zona_adjudicacion_choices': zona_adjudicacion_choices,
    }

    return render(request, 'pagos/editar_capacity.html', context)


# Vista para eliminar un registro del modelo Capacity
@login_required  # Requiere que el usuario esté autenticado
def eliminar_capacity(request, pk):
    # Obtiene el registro de Capacity correspondiente al ID (pk) o lanza un error 404 si no se encuentra
    record = get_object_or_404(Capacity, pk=pk)

    # Si la solicitud es de tipo POST (el usuario ha confirmado la eliminación)
    if request.method == 'POST':
        # Elimina el registro de la base de datos
        record.delete()
        # Muestra un mensaje de éxito al usuario indicando que el registro ha sido eliminado
        messages.success(request, 'El registro ha sido eliminado con éxito.')
        # Redirige a la vista 'ver_datos_capacity_mes' mostrando los datos del mes y año correspondiente
        return redirect('ver_datos_capacity_mes', year=record.fecha_extract_registro.year, month=record.fecha_extract_registro.month)

    # Si la solicitud es de tipo GET, se muestra una página de confirmación de eliminación
    return render(request, 'pagos/eliminar_capacity.html', {'record': record})


# Vista para crear un nuevo registro en el modelo Capacity

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .models import Capacity
from .forms import CapacityForm

@login_required
def crear_capacity(request):
    if request.method == 'POST':
        form = CapacityForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('dashboard_capacity')
    else:
        form = CapacityForm()

    # Obtener valores únicos para los campos desplegables
    nombre_team_choices = Capacity.objects.values_list('nombre_team', flat=True).distinct()
    especialidad_usuario_choices = Capacity.objects.values_list('especialidad_usuario', flat=True).distinct()
    area_choices = Capacity.objects.values_list('area', flat=True).distinct()
    zona_operacional_choices = Capacity.objects.values_list('zona_operacional', flat=True).distinct()
    especialidad_team_choices = Capacity.objects.values_list('especialidad_team', flat=True).distinct()
    zona_adjudicacion_choices = Capacity.objects.values_list('zona_adjudicacion', flat=True).distinct()

    context = {
        'form': form,
        'nombre_team_choices': nombre_team_choices,
        'especialidad_usuario_choices': especialidad_usuario_choices,
        'area_choices': area_choices,
        'zona_operacional_choices': zona_operacional_choices,
        'especialidad_team_choices': especialidad_team_choices,
        'zona_adjudicacion_choices': zona_adjudicacion_choices,
    }

    return render(request, 'pagos/crear_capacity.html', context)

from django.db.models import Count
# Función para eliminar registros duplicados en el modelo Capacity
@login_required
def eliminar_duplicados_capacity():
    try:
        # Encontrar los registros duplicados en Capacity basados en los campos: 'nombre_tecnico', 'rut_tecnico', 
        # 'especialidad_usuario', y 'zona_adjudicacion'. Se agrupan estos campos y se cuenta cuántas veces se repiten.
        duplicates = (
            Capacity.objects.values('nombre_tecnico', 'rut_tecnico', 'especialidad_usuario', 'zona_adjudicacion')
            .annotate(count=Count('id'))  # Cuenta la cantidad de registros por cada combinación de los campos especificados
            .filter(count__gt=1)  # Filtra para obtener solo aquellos grupos que tienen más de un registro (es decir, duplicados)
        )

        # Iterar sobre cada grupo de duplicados
        for duplicate in duplicates:
            # Obtener todos los registros duplicados que coinciden con los valores de 'nombre_tecnico', 'rut_tecnico',
            # 'especialidad_usuario', y 'zona_adjudicacion', ordenándolos por 'id'
            duplicate_records = Capacity.objects.filter(
                nombre_tecnico=duplicate['nombre_tecnico'],
                rut_tecnico=duplicate['rut_tecnico'],
                especialidad_usuario=duplicate['especialidad_usuario'],
                zona_adjudicacion=duplicate['zona_adjudicacion'],
            ).order_by('id')[1:]  # Selecciona todos los registros duplicados excepto el primero (que se mantendrá)

            # Iterar sobre los registros duplicados seleccionados y eliminarlos uno por uno
            for record in duplicate_records:
                record.delete()

        # Imprimir mensaje de éxito después de eliminar todos los duplicados
        print("Duplicados eliminados correctamente.")
    except Exception as e:
        # Si ocurre un error, imprimir el mensaje de error
        print(f"Error al eliminar duplicados: {e}")

from django.http import HttpResponse
import pandas as pd
from .models import NuevaTablaPagos

# Vista para descargar los datos de la tabla NuevaTablaPagos en formato Excel
@login_required
def descargar_nueva_tabla_pagos(request):
    # Obtiene todos los registros de la tabla NuevaTablaPagos con los campos especificados
    datos = NuevaTablaPagos.objects.all().values(
        'sociedad', 'material', 'especialidad', 'zona_adjudicacion', 
        'zona_operacional', 'nuevo_precio', 'desvio_porcentaje', 'area', 'anio'
    )
    
    # Convierte los datos obtenidos a un DataFrame de pandas
    df = pd.DataFrame(list(datos))
    
    # Renombra la columna 'anio' a 'año'
    df.rename(columns={'anio': 'año'}, inplace=True)

    # Crea una respuesta HTTP con el tipo de contenido 'application/vnd.ms-excel'
    response = HttpResponse(content_type='application/vnd.ms-excel')
    # Define los encabezados para la descarga del archivo con un nombre específico
    response['Content-Disposition'] = 'attachment; filename="nueva_tabla_pagos.xlsx"'

    # Guarda el DataFrame en el archivo Excel sin incluir los índices de pandas (index=False)
    df.to_excel(response, index=False)

    # Retorna la respuesta para que el archivo se descargue
    return response


# Vista para ver todos los registros de Capacity con paginación
@login_required  # Requiere que el usuario esté autenticado
def ver_datos_capacity(request):
    # Obtiene todos los registros del modelo Capacity y los ordena por 'id' de manera ascendente
    datos_list = Capacity.objects.all().order_by('id')  # Ordenar por 'id'
    
    # Paginación: divide los registros en páginas de 10 elementos cada una
    paginator = Paginator(datos_list, 10)  # 10 registros por página
    # Obtiene el número de página actual desde los parámetros GET
    page = request.GET.get('page')
    
    try:
        # Si se proporciona un número de página válido, obtiene los datos de esa página
        datos = paginator.page(page)
    except PageNotAnInteger:
        # Si el número de página no es un entero, muestra la primera página
        datos = paginator.page(1)
    except EmptyPage:
        # Si el número de página es mayor que el número total de páginas, muestra la última página
        datos = paginator.page(paginator.num_pages)

    # Define el contexto a pasar a la plantilla, con los registros paginados
    context = {
        'datos': datos  # Los registros de Capacity paginados
    }

    # Renderiza la plantilla 'ver_datos_capacity.html', pasando los datos paginados en el contexto
    return render(request, 'pagos/ver_datos_capacity.html', context)

# Vista para ver todos los registros de NuevaTablaPagos con paginación
@login_required  # Requiere que el usuario esté autenticado
def ver_nuevatabladepagos(request):
    # Obtiene todos los registros del modelo NuevaTablaPagos y los ordena por 'id'
    datos_list = NuevaTablaPagos.objects.all().order_by('id')  # Ordenar por 'id'
    
    # Paginación: divide los registros en páginas de 10 elementos cada una
    paginator = Paginator(datos_list, 10)  # 10 registros por página
    # Obtiene el número de página actual desde los parámetros GET
    page = request.GET.get('page')
    
    try:
        # Si se proporciona un número de página válido, obtiene los datos de esa página
        datos = paginator.page(page)
    except PageNotAnInteger:
        # Si el número de página no es un entero, muestra la primera página
        datos = paginator.page(1)
    except EmptyPage:
        # Si el número de página es mayor que el número total de páginas, muestra la última página
        datos = paginator.page(paginator.num_pages)

    # Define el contexto a pasar a la plantilla, con los registros paginados
    context = {
        'datos': datos  # Los registros de NuevaTablaPagos paginados
    }

    # Renderiza la plantilla 'ver_nueva_tabla_pagos.html', pasando los datos paginados en el contexto
    return render(request, 'pagos/ver_nueva_tabla_pagos.html', context)

@login_required
def descargar_excel_pago_capacity(request, year, month):
    try:
        # Obtener los registros de Capacity para el mes y año especificado
        capacities = Capacity.objects.filter(
            fecha_extract_registro__year=year,
            fecha_extract_registro__month=month
        )

        if not capacities.exists():
            return HttpResponse("No hay registros para este mes.")
        def obtener_proyecto(nombre_team):
            # Reglas para asignar proyectos basados en el valor de nombre_team
            if 'COMANDO' in nombre_team:
                return 'COMANDO'
            elif 'ONNET' in nombre_team:
                return 'ONNET'
            elif 'FON' in nombre_team:
                return 'FON'
            elif 'PROYECTO DESIGENIA' in nombre_team:
                return 'PROYECTO DESIGENIA'
            elif 'SAE' in nombre_team:
                return 'SAE'
            # Añadir más reglas según sea necesario
            return 'COMANDO'

        # Convertir Capacity y NuevaTablaPagos a DataFrames
        df_capacity = pd.DataFrame(list(capacities.values()))
        df_nueva_tabla_pagos = pd.DataFrame(list(NuevaTablaPagos.objects.all().values()))

        # Limpiar y normalizar los datos
        df_capacity['especialidad_usuario'] = df_capacity['especialidad_usuario'].fillna('').astype(str).str.strip().str.upper()
        df_nueva_tabla_pagos['especialidad'] = df_nueva_tabla_pagos['especialidad'].fillna('').astype(str).str.strip().str.upper()

        # Asegurarse de que las zonas y áreas están correctamente formateadas
        df_capacity['zona_operacional'] = df_capacity['zona_operacional'].fillna('').astype(str).str.strip().str.upper()
        df_nueva_tabla_pagos['zona_operacional'] = df_nueva_tabla_pagos['zona_operacional'].fillna('').astype(str).str.strip().str.upper()

        df_capacity['zona_adjudicacion'] = df_capacity['zona_adjudicacion'].fillna('').astype(str).str.strip().str.upper()
        df_nueva_tabla_pagos['zona_adjudicacion'] = df_nueva_tabla_pagos['zona_adjudicacion'].fillna('').astype(str).str.strip().str.upper()

        df_capacity['area'] = df_capacity['area'].fillna('').astype(str).str.strip().str.upper()
        df_nueva_tabla_pagos['area'] = df_nueva_tabla_pagos['area'].fillna('').astype(str).str.strip().str.upper()

        # Crear la columna 'proyecto' en df_capacity, si no existe
        df_capacity['proyecto'] = df_capacity['nombre_team'].apply(obtener_proyecto)

        # Asignar el tipo de red basado en la especialidad
        def asignar_tipo_red(especialidad_usuario):
            especialidades_tipo_red = {
                'FIJA': ["(F)", "FIJA", "LINEA - CANALIZACIÓN", "LOCALIZADOR - EMPALMADOR DE FIBRA", 
                        "JEFE ESPECIALISTA DE ATENCIÓN EN CAMPO (M/F)", "SERVICIOS PRIVADOS", 
                        "LOCALIZADOR - EMPALMADOR EXPERTO COBRE / FIBRA - VERSION PLUS",'LOCALIZADOR – EMPALMADOR EXPERTO COBRE / FIBRA - VERSION PLUS'],
                'MOVIL': ["(M)", "MÓVIL", "CLIMA", "RADIOFRECUENCIA", "ENERGIA",
                        "OPERACIÓN MOVIL SERVICIO MULTISKILL N° 1", "OPERACIÓN MOVIL SERVICIO MULTISKILL N° 2",
                        "OPERACIÓN MOVIL SERV SUPERVISOR DE CAMPO", "OPERACIÓN MOVIL SERV VEHICULO PATNER",
                        "OPERACIÓN MOVIL SERVICIO JEFE ESPECIALISTA", "OPERACIÓN MOVIL SERVICIO ENERGÍA"]
            }
            for tipo, especialidades in especialidades_tipo_red.items():
                if any(esp in especialidad_usuario for esp in especialidades):
                    return tipo
            return "desconocido"

        # Función para clasificar si es un vehículo o una especialidad
        def clasificar_tipo(especialidad_usuario):
            """
            Clasifica un valor de especialidad en 'ESPECIALIDAD' o 'VEHICULO' 
            según las palabras clave asociadas.
            """
            # Lista de palabras clave que indican que es un vehículo
            palabras_vehiculo = [
                "VEHICULO", "PATNER", "CAMIONETA", "SERV VEHICULO", 
                "VEHICULO PATNER", "VEHICULO","CAMIÓN 3/4 (F)","CAMIÓN 3/4 (M)"
            ]

            # Verificar si la especialidad contiene alguna de las palabras clave de 'VEHICULO'
            if any(palabra in especialidad_usuario for palabra in palabras_vehiculo):
                return 'VEHICULO'

            # Si no contiene ninguna palabra de VEHICULO, lo clasificamos como 'ESPECIALIDAD'
            return 'ESPECIALIDAD'

        # Asignar el tipo de red basado en la especialidad
        df_capacity['tipo_red'] = df_capacity['especialidad_usuario'].apply(asignar_tipo_red)

        # Crear columna 'tipo_especialidad'
        df_capacity['tipo_especialidad'] = df_capacity['especialidad_usuario'].apply(clasificar_tipo)

        # Fusión de Capacity y NuevaTablaPagos utilizando un 'left join'
        df_merged = pd.merge(
            df_capacity,
            df_nueva_tabla_pagos,
            left_on=['especialidad_usuario', 'zona_adjudicacion', 'zona_operacional', 'area'],
            right_on=['especialidad', 'zona_adjudicacion', 'zona_operacional', 'area'],
            how='left'
        )

        # Eliminar duplicados basados en las claves necesarias
        df_merged = df_merged.drop_duplicates(subset=['nombre_tecnico', 'rut_tecnico', 'especialidad_usuario', 'zona_adjudicacion'])

        # Asegurarse de que las columnas numéricas están correctas, reemplazar NaN con 0
        df_merged['nuevo_precio'] = pd.to_numeric(df_merged['nuevo_precio'], errors='coerce').fillna(0)
        df_merged['total_hh_phi'] = pd.to_numeric(df_merged['total_hh_phi'], errors='coerce').fillna(0)

        # Calcular Q con 9 decimales usando Decimal, pero luego convertirlo a float
        df_merged['Q'] = df_merged.apply(
            lambda row: float((Decimal(row['total_hh_phi']) / Decimal(176)).quantize(Decimal('0.000000001'), rounding=ROUND_HALF_UP)),
            axis=1
        )

        # Calcular Total_a_Pago, manejando casos donde nuevo_precio_2024 o Q podrían ser NaN
        df_merged['Total_a_Pago'] = df_merged.apply(
            lambda row: (row['nuevo_precio'] * row['Q'])
            if pd.notna(row['nuevo_precio']) and pd.notna(row['Q']) 
            else 0, 
            axis=1
        ).astype(float)

        # Filtrar las columnas requeridas
        columnas_requeridas = [
            'nombre_tecnico', 'nombre_team', 'especialidad_usuario', 'rut_tecnico', 
            'zona_operacional', 'zona_adjudicacion', 'total_hh', 'phi', 
            'total_hh_phi', 'nuevo_precio', 'Q', 'Total_a_Pago', 
            'tipo_red', 'proyecto', 'tipo_especialidad', 'area'
        ]
        
        df_resumido = df_merged[columnas_requeridas]

        # Crear el archivo Excel en memoria
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df_resumido.to_excel(writer, index=False, sheet_name='PagoFinalCapacity')

            # Obtener el workbook y worksheet para aplicar formateo
            workbook = writer.book
            worksheet = writer.sheets['PagoFinalCapacity']

            # Formato de número con dos decimales en nuevo_precio_2024, sin decimales en Total_a_Pago y nueve en Q
            formato_numero_dos_decimales = workbook.add_format({'num_format': '#,##0.00'})
            formato_numero_nueve_decimales = workbook.add_format({'num_format': '#,##0.000000000'})
            formato_numero_sin_decimales = workbook.add_format({'num_format': '#,##0'})  # Sin decimales para Total_a_Pago

            # Aplicar el formato a las columnas correspondientes
            worksheet.set_column('J:J', None, formato_numero_dos_decimales)  # Columna 'nuevo_precio_2024'
            worksheet.set_column('K:K', None, formato_numero_nueve_decimales)  # Columna 'Q' con 9 decimales como número
            worksheet.set_column('L:L', None, formato_numero_sin_decimales)  # Columna 'Total_a_Pago' sin decimales

            writer.close()

        # Configurar la respuesta HTTP para la descarga del archivo Excel
        output.seek(0)
        response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=PagoFinalCapacity_{year}_{month}.xlsx'

        return response

    except Exception as e:
        logger.error(f"Error al descargar el archivo Excel: {str(e)}")
        return HttpResponse(f"Error al descargar el archivo Excel: {str(e)}")


from decimal import Decimal
from .models import TablaDePagoCapacity
from decimal import Decimal
from datetime import datetime
from django.shortcuts import render
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from .models import TablaDePagoCapacity, Capacity, NuevaTablaPagos
import pandas as pd
from decimal import ROUND_HALF_UP
import logging

logger = logging.getLogger(__name__)

@login_required
def generar_tabla_pago_capacity_mes(request, year, month):
    try:
        # Obtener los registros de Capacity para el mes y año especificado
        capacities = Capacity.objects.filter(
            fecha_extract_registro__year=year,
            fecha_extract_registro__month=month
        )

        if not capacities.exists():
            return HttpResponse("No hay registros para este mes.")
        
        # Función de obtención de proyecto según nombre_team
        def obtener_proyecto(nombre_team):
            if 'COMANDO' in nombre_team:
                return 'COMANDO'
            elif 'ONNET' in nombre_team:
                return 'ONNET'
            elif 'FON' in nombre_team:
                return 'FON'
            elif 'PROYECTO DESIGENIA' in nombre_team:
                return 'PROYECTO DESIGENIA'
            elif 'SAE' in nombre_team:
                return 'SAE'
            return 'COMANDO'
        
        # Convertir Capacity y NuevaTablaPagos a DataFrames
        df_capacity = pd.DataFrame(list(capacities.values()))
        df_nueva_tabla_pagos = pd.DataFrame(list(NuevaTablaPagos.objects.all().values()))

        # Normalización de datos
        df_capacity['especialidad_usuario'] = df_capacity['especialidad_usuario'].fillna('').astype(str).str.strip().str.upper()
        df_nueva_tabla_pagos['especialidad'] = df_nueva_tabla_pagos['especialidad'].fillna('').astype(str).str.strip().str.upper()

        df_capacity['zona_operacional'] = df_capacity['zona_operacional'].fillna('').astype(str).str.strip().str.upper()
        df_nueva_tabla_pagos['zona_operacional'] = df_nueva_tabla_pagos['zona_operacional'].fillna('').astype(str).str.strip().str.upper()

        df_capacity['zona_adjudicacion'] = df_capacity['zona_adjudicacion'].fillna('').astype(str).str.strip().str.upper()
        df_nueva_tabla_pagos['zona_adjudicacion'] = df_nueva_tabla_pagos['zona_adjudicacion'].fillna('').astype(str).str.strip().str.upper()

        df_capacity['area'] = df_capacity['area'].fillna('').astype(str).str.strip().str.upper()
        df_nueva_tabla_pagos['area'] = df_nueva_tabla_pagos['area'].fillna('').astype(str).str.strip().str.upper()

        # Crear la columna 'proyecto' en df_capacity
        df_capacity['proyecto'] = df_capacity['nombre_team'].apply(obtener_proyecto)

        # Asignar el tipo de red basado en la especialidad
        def asignar_tipo_red(especialidad_usuario):
            especialidades_tipo_red = {
                'FIJA': ["(F)", "FIJA", "LINEA - CANALIZACIÓN", "LOCALIZADOR - EMPALMADOR DE FIBRA", 
                        "JEFE ESPECIALISTA DE ATENCIÓN EN CAMPO (M/F)", "SERVICIOS PRIVADOS", 
                        "LOCALIZADOR - EMPALMADOR EXPERTO COBRE / FIBRA - VERSION PLUS", 'LOCALIZADOR – EMPALMADOR EXPERTO COBRE / FIBRA - VERSION PLUS'],
                'MOVIL': ["(M)", "MÓVIL", "CLIMA", "RADIOFRECUENCIA", "ENERGIA",
                        "OPERACIÓN MOVIL SERVICIO MULTISKILL N° 1", "OPERACIÓN MOVIL SERVICIO MULTISKILL N° 2",
                        "OPERACIÓN MOVIL SERV SUPERVISOR DE CAMPO", "OPERACIÓN MOVIL SERV VEHICULO PATNER",
                        "OPERACIÓN MOVIL SERVICIO JEFE ESPECIALISTA", "OPERACIÓN MOVIL SERVICIO ENERGÍA"]
            }
            for tipo, especialidades in especialidades_tipo_red.items():
                if any(esp in especialidad_usuario for esp in especialidades):
                    return tipo
            return "desconocido"

        # Función para clasificar si es un vehículo o una especialidad
        # Función mejorada para clasificar entre ESPECIALIDAD y VEHICULO
        def clasificar_tipo(especialidad_usuario):
            """
            Clasifica un valor de especialidad en 'ESPECIALIDAD' o 'VEHICULO' 
            según las palabras clave asociadas.
            """
            # Lista de palabras clave que indican que es un vehículo
            palabras_vehiculo = [
                "VEHICULO", "PATNER", "CAMIONETA", "SERV VEHICULO", 
                "VEHICULO PATNER", "VEHICULO","CAMIÓN 3/4 (F)","CAMIÓN 3/4 (M)"
            ]

            # Verificar si la especialidad contiene alguna de las palabras clave de 'VEHICULO'
            if any(palabra in especialidad_usuario for palabra in palabras_vehiculo):
                return 'VEHICULO'

            # Si no contiene ninguna palabra de VEHICULO, lo clasificamos como 'ESPECIALIDAD'
            return 'ESPECIALIDAD'
        

        # Asignar el tipo de red y tipo de especialidad
        df_capacity['tipo_red'] = df_capacity['especialidad_usuario'].apply(asignar_tipo_red)
        df_capacity['tipo_especialidad'] = df_capacity['especialidad_usuario'].apply(clasificar_tipo)

        # Fusión de Capacity y NuevaTablaPagos utilizando un 'left join'
        df_merged = pd.merge(
            df_capacity,
            df_nueva_tabla_pagos,
            left_on=['especialidad_usuario', 'zona_adjudicacion', 'zona_operacional', 'area'],
            right_on=['especialidad', 'zona_adjudicacion', 'zona_operacional', 'area'],
            how='left'
        )

        # Eliminar duplicados basados en las claves necesarias
        df_merged = df_merged.drop_duplicates(subset=['nombre_tecnico', 'rut_tecnico', 'especialidad_usuario', 'zona_adjudicacion'])

        # Asegurarse de que las columnas numéricas están correctas, reemplazar NaN con 0
        df_merged['nuevo_precio'] = pd.to_numeric(df_merged['nuevo_precio'], errors='coerce').fillna(Decimal('0'))
        df_merged['total_hh_phi'] = pd.to_numeric(df_merged['total_hh_phi'], errors='coerce').fillna(Decimal('0'))

        # Calcular Q con 9 decimales usando Decimal, pero luego convertirlo a float
        df_merged['Q'] = df_merged.apply(
            lambda row: float((Decimal(row['total_hh_phi']) / Decimal(176)).quantize(Decimal('0.000000001'), rounding=ROUND_HALF_UP)),
            axis=1
        )

        # Calcular Total_a_Pago, manejando casos donde nuevo_precio o Q podrían ser NaN
        df_merged['Total_a_Pago'] = df_merged.apply(
            lambda row: (row['nuevo_precio'] * row['Q']) if pd.notna(row['nuevo_precio']) and pd.notna(row['Q']) else Decimal('0'),
            axis=1
        ).astype(float)

        # Verificar si las columnas 'Q' y 'Total_a_Pago' están presentes en df_merged
        if 'Q' not in df_merged.columns or 'Total_a_Pago' not in df_merged.columns:
            return HttpResponse("Error: No se encontraron las columnas Q o Total_a_Pago en el DataFrame.")

        # Filtrar las columnas requeridas y convertir a una lista de diccionarios
        columnas_requeridas = [
            'nombre_tecnico', 'nombre_team', 'especialidad_usuario', 'rut_tecnico',
            'zona_operacional', 'zona_adjudicacion', 'total_hh', 'phi',
            'total_hh_phi', 'nuevo_precio', 'Q', 'Total_a_Pago',
            'tipo_red', 'proyecto', 'tipo_especialidad', 'area'
        ]

        # Verificar que las columnas estén presentes en el DataFrame antes de filtrarlas
        if not all(col in df_merged.columns for col in columnas_requeridas):
            return HttpResponse("Error: Faltan columnas en el DataFrame.")

        # Filtrar las columnas requeridas y convertir a una lista de diccionarios
        df_resumido = df_merged[columnas_requeridas].fillna(Decimal('0'))
        datos = df_resumido.to_dict(orient='records')  # Convierte a una lista de diccionarios

        # Guardar cada fila en la base de datos
        # Guardar cada fila en la base de datos
        for _, row in df_resumido.iterrows():
            # Verificar que la columna 'fecha' está presente y tiene un valor
            if 'fecha' not in row or row['fecha'] is None:
                logger.warning(f"Fecha no presente o vacía para: {row['nombre_tecnico']} - {row['especialidad_usuario']} - {row['zona_adjudicacion']}")
                continue  # Si no hay fecha, continuar con el siguiente registro
            
            # Verificar si ya existe un registro con los mismos campos (pero fecha distinta)
            capacidad_existente = TablaDePagoCapacity.objects.filter(
                nombre_tecnico=row['nombre_tecnico'],
                especialidad_usuario=row['especialidad_usuario'],
                zona_adjudicacion=row['zona_adjudicacion'],
                fecha=row['fecha']
            ).first()

            # Si no existe, crear un nuevo registro
            if capacidad_existente is None:
                TablaDePagoCapacity.objects.create(
                    nombre_tecnico=row['nombre_tecnico'],
                    rut_tecnico=row['rut_tecnico'],
                    especialidad_usuario=row['especialidad_usuario'],
                    zona_adjudicacion=row['zona_adjudicacion'],
                    nombre_team=row['nombre_team'],
                    zona_operacional=row['zona_operacional'],
                    total_hh=row['total_hh'],
                    phi=row['phi'],
                    total_hh_phi=row['total_hh_phi'],
                    nuevo_precio=row['nuevo_precio'],
                    Q=row['Q'],
                    Total_a_Pago=row['Total_a_Pago'],
                    tipo_red=row['tipo_red'],
                    proyecto=row['proyecto'],
                    tipo_especialidad=row['tipo_especialidad'],
                    area=row['area'],
                    fecha=row['fecha']  # Aseguramos que la fecha se asigne correctamente
                )
            else:
                # Si ya existe, actualizar el registro existente
                capacidad_existente.nombre_team = row['nombre_team']
                capacidad_existente.zona_operacional = row['zona_operacional']
                capacidad_existente.total_hh = row['total_hh']
                capacidad_existente.phi = row['phi']
                capacidad_existente.total_hh_phi = row['total_hh_phi']
                capacidad_existente.nuevo_precio = row['nuevo_precio']
                capacidad_existente.Q = row['Q']
                capacidad_existente.Total_a_Pago = row['Total_a_Pago']
                capacidad_existente.tipo_red = row['tipo_red']
                capacidad_existente.proyecto = row['proyecto']
                capacidad_existente.tipo_especialidad = row['tipo_especialidad']
                capacidad_existente.area = row['area']
                capacidad_existente.fecha = row['fecha']  # Aseguramos que la fecha se asigne correctamente

                # Guardar el registro actualizado
                capacidad_existente.save()





            # Si el registro es un vehículo, verificar condiciones adicionales
            if row['tipo_especialidad'] == 'VEHICULO' and fecha is None:
                logger.warning(f"Registro clasificado como vehículo sin fecha: {row}")
                # Asignar una fecha predeterminada para vehículos si es necesario
                fecha = datetime(year, month, 1)  # Ejemplo: primera fecha del mes procesado

        # Renderizar la tabla en la plantilla
        return render(request, 'pagos/tabla_pago_capacity_mes.html', {
            'datos': datos,
            'year': year,
            'month': month
        })

    except Exception as e:
        logger.error(f"Error al generar la tabla de pago: {str(e)}")
        return HttpResponse(f"Error al generar la tabla de pago: {str(e)}")

# Mapeo de especialidades
def mapear_especialidad(especialidad_usuario):
    # Mapeo manual de especialidades para garantizar coincidencias exactas
    mapeo_especialidades = {
        "MULTIEXPERTO RED FIBRA (F)": "OPERACIÓN FIJA MULTIEXPERTO RED FIBRA",
        "VEHICULO CAMIONETA 4X4 (M)": "OPERACIÓN MOVIL SERV VEHICULO CAMIONETA 4X4",
        # Agregar otros mapeos necesarios
    }
    return mapeo_especialidades.get(especialidad_usuario, especialidad_usuario)

# Función para asignar el proyecto en base a nombre_team
def obtener_proyecto(nombre_team):
    proyectos = {
        'ATP': 'ATP',
        'CODELCO': 'CODELCO',
        'COMANDO': 'COMANDO',
        # Otros proyectos según palabras clave en nombre_team
    }
    for keyword, proyecto in proyectos.items():
        if keyword in nombre_team:
            return proyecto
    return "PROYECTO_DEFAULT"

# Mapeo de especialidades
def mapear_especialidad(especialidad_usuario):
    # Mapeo manual de especialidades para garantizar coincidencias exactas
    mapeo_especialidades = {
        "VEHICULO CAMIONETA 4X4 (M)": "VEHICULO CAMIONETA 4X4 (M)",  # Verifica si la base de datos usa este formato
        "VEHICULO CAMIONETA 4X4 (F)": "VEHICULO CAMIONETA 4X4 (F)",  # Igual para vehículos femeninos
        "SUPERVISOR DE CAMPO (M)": "SUPERVISOR DE CAMPO (M)",
        "JEFE ESPECIALISTA (F)": "JEFE ESPECIALISTA (F)",
        "VEHICULO PATNER (M)": "VEHICULO PATNER (M)",
        # Agregar otros mapeos necesarios
    }
    return mapeo_especialidades.get(especialidad_usuario, especialidad_usuario)



# Vista para ver los registros de SLAReport con paginación
def ver_sla_report(request):
    # Obtiene todos los registros del modelo SLAReport y los ordena por 'id' (puedes ajustar el orden si lo deseas)
    sla_list = SLAReport.objects.all().order_by('id')  # Ordenar por 'id'
    
    # Paginación: divide los registros en páginas de 10 elementos cada una
    paginator = Paginator(sla_list, 10)  # 10 registros por página
    
    # Obtiene el número de página actual desde los parámetros GET
    page_number = request.GET.get('page')
    
    # Obtiene el conjunto de objetos correspondientes a la página actual
    page_obj = paginator.get_page(page_number)

    # Renderiza la plantilla 'ver_sla_report.html', pasando los registros paginados
    return render(request, 'pagos/ver_sla_report.html', {'page_obj': page_obj})


def resumen_pagos_sla(request):
    # Lógica para obtener el resumen de pagos
    context = {
        # Datos que necesites pasar a la plantilla
    }
    return render(request, 'tu_template_resumen.html', context)


def generar_tabla_pago_mes(request, year, month):
    # Lógica para generar la tabla de pago
    context = {
        'year': year,
        'month': month,
        # otros datos que necesites
    }
    return render(request, 'tu_template.html', context)

# Vista para editar un registro de SLAReport
from django.shortcuts import render, get_object_or_404, redirect
from .models import SLAReport
from .forms import SLAReportForm

def editar_sla_report(request, pk):
    # Obtiene el registro de SLAReport correspondiente al ID (pk) o lanza un error 404 si no se encuentra
    sla_report = get_object_or_404(SLAReport, pk=pk)

    # Obtén las zonas de adjudicación, zonas operativas y proyectos únicos desde la base de datos
    zonas_adjudicacion = SLAReport.objects.values('zona_adjudicacion').distinct()
    zonas_operativas = SLAReport.objects.values('zona_operativa').distinct()
    proyectos = SLAReport.objects.values('proyecto').distinct()

    # Si la solicitud es de tipo POST (el formulario ha sido enviado)
    if request.method == 'POST':
        # Crea una instancia del formulario con los datos enviados y asocia el registro existente (instance=sla_report)
        form = SLAReportForm(request.POST, instance=sla_report)
        # Verifica si el formulario es válido (pasa todas las validaciones)
        if form.is_valid():
            # Guarda los cambios en el registro existente en la base de datos
            form.save()
            # Redirige a la vista 'ver_sla_report' después de guardar los cambios
            return redirect('ver_sla_report')
    else:
        # Si el método no es POST (es decir, es GET), muestra el formulario con los datos actuales del registro
        form = SLAReportForm(instance=sla_report)

    # Renderiza la plantilla 'editar_sla_report.html', pasando el formulario y los datos dinámicos al contexto
    context = {
        'form': form,
        'zonas_adjudicacion': zonas_adjudicacion,
        'zonas_operativas': zonas_operativas,
        'proyectos': proyectos,
    }
    return render(request, 'pagos/editar_sla_report.html', context)


def eliminar_sla_report(request, pk):
    try:
        sla_report = SLAReport.objects.get(pk=pk)
        sla_report.delete()
        messages.success(request, 'SLA Report eliminado exitosamente.')
    except SLAReport.DoesNotExist:
        messages.error(request, 'El SLA Report no existe.')
    except Exception as e:
        messages.error(request, f'Error al eliminar el SLA Report: {str(e)}')
    
    return redirect('ver_sla_report')



# Vista para crear un nuevo registro en el modelo SLAReport
from pagos.models import ZonaOperativa, ZonaAdjudicacion

from django.shortcuts import render
from pagos.models import SLAReport

from django.shortcuts import render
from pagos.models import SLAReport

def crear_sla_report(request):
    # Obtener zonas adjudicación únicas
    zonas_adjudicacion = SLAReport.objects.values('zona_adjudicacion').distinct()

    # Obtener zonas operativas únicas
    zonas_operativas = SLAReport.objects.values('zona_operativa').distinct()

    # Obtener proyectos únicos
    proyectos = SLAReport.objects.values('proyecto').distinct()

    context = {
        'zonas_adjudicacion': zonas_adjudicacion,
        'zonas_operativas': zonas_operativas,
        'proyectos': proyectos,
    }
    return render(request, 'pagos/crear_sla_report.html', context)


def ver_resumen_pagos(request):
    # Agrupar por zona_operacional y sumar los valores de las columnas correspondientes
    pagos = TablaDePagoCapacity.objects.values('zona_operacional').annotate(
        total_especialidad=Coalesce(Sum(F('total_hh_phi'), output_field=DecimalField()), Value(0, output_field=DecimalField())),
        total_vehiculo=Coalesce(Sum(F('phi'), output_field=DecimalField()), Value(0, output_field=DecimalField())),
        total_gastos_admin=Coalesce(Sum(F('total_hh'), output_field=DecimalField()), Value(0, output_field=DecimalField())),
        total_lpu=Coalesce(Sum(F('nuevo_precio'), output_field=DecimalField()), Value(0, output_field=DecimalField())),  # Cambiado a nuevo_precio
        total_sla=Coalesce(Sum(F('Q'), output_field=DecimalField()), Value(0, output_field=DecimalField())),
        total_general=Coalesce(Sum(F('Total_a_Pago'), output_field=DecimalField()), Value(0, output_field=DecimalField())),
        # Contamos el número de especialidades y vehículos
        count_especialidades=Count('especialidad_usuario', distinct=True),
        count_vehiculos=Count('tipo_especialidad', filter=F('tipo_especialidad') == 'vehículo', distinct=True)
    )

    contexto = {
        'pagos': pagos,
    }

    return render(request, 'pagos/resumen_pagos.html', contexto)


from collections import defaultdict
from django.db.models import Sum
from datetime import date
from django.shortcuts import render
from django.contrib.auth.decorators import login_required

@login_required
def ver_resumen_mensual(request, month, year):
    try:
        # Conversión y validación de mes y año
        month = int(month)
        year = int(year)

        if month < 1 or month > 12:
            return render(request, 'pagos/error.html', {'mensaje': 'Mes inválido.'})

        # Calcular las fechas de inicio y fin del mes
        inicio_mes = date(year, month, 1)
        fin_mes = date(year, month + 1, 1) if month != 12 else date(year + 1, 1, 1)

        # Filtrar registros del mes
        registros = TablaDePagoCapacity.objects.filter(fecha__gte=inicio_mes, fecha__lt=fin_mes)

        # Agregaciones globales (Total_a_Pago, total_hh, Q)
        agregados = registros.aggregate(
            total_pago=Sum('Total_a_Pago'),
            total_hh=Sum('total_hh'),
            total_q=Sum('Q')
        )

        # Crear el diccionario resumen global
        resumen = {
            'total_pago': agregados['total_pago'] or 0,
            'total_hh': agregados['total_hh'] or 0,
            'total_q': agregados['total_q'] or 0
        }

        # Agregación por área
        resumen_areas = registros.values('area').annotate(suma_pago=Sum('Total_a_Pago'))
        total_general_area = sum((r['suma_pago'] or 0) for r in resumen_areas)

        # Áreas fijas
        todas_las_areas = ["LARI", "ZENER", "COBRA"]
        dict_areas = {item['area']: item['suma_pago'] for item in resumen_areas}
        resumen_areas_final = []
        for area in todas_las_areas:
            resumen_areas_final.append({
                'area': area,
                'suma_pago': dict_areas.get(area, 0)
            })

        # Agrupación por zona_operacional, tipo_especialidad y proyecto
        resumen_zonas = registros.values('zona_operacional', 'tipo_especialidad', 'proyecto').annotate(
            total_pago=Sum('Total_a_Pago'),
            total_hh=Sum('total_hh'),
            total_q=Sum('Q')
        )

        # Diccionario para los totales por zona
        resumen_por_zona = {}
        resumen_por_proyecto = defaultdict(dict)

        proyectos_encontrados = set()
        zonas_encontradas = set()

        # Recorrer datos agregados por zona
        for zona_data in resumen_zonas:
            zona_nombre = zona_data.get('zona_operacional', '')
            tipo_especialidad = zona_data.get('tipo_especialidad', '')
            proyecto = zona_data.get('proyecto', '')
            total_pago_zona = zona_data['total_pago'] or 0

            if proyecto:
                proyectos_encontrados.add(proyecto)
            if zona_nombre:
                zonas_encontradas.add(zona_nombre)

            if zona_nombre and zona_nombre not in resumen_por_zona:
                resumen_por_zona[zona_nombre] = {
                    'total_especialidad': 0,
                    'total_vehiculo': 0
                }

            # Sumar a especialidad o vehículo
            if tipo_especialidad == "ESPECIALIDAD":
                resumen_por_zona[zona_nombre]['total_especialidad'] += total_pago_zona
            elif tipo_especialidad == "VEHICULO":
                resumen_por_zona[zona_nombre]['total_vehiculo'] += total_pago_zona

            # Sumar al proyecto por zona
            if proyecto and zona_nombre:
                if zona_nombre not in resumen_por_proyecto[proyecto]:
                    resumen_por_proyecto[proyecto][zona_nombre] = 0
                resumen_por_proyecto[proyecto][zona_nombre] += total_pago_zona

        proyectos = sorted(proyectos_encontrados)
        todas_las_zonas = sorted(zonas_encontradas)

        # Construir la lista final para la tabla de proyectos por zona
        resumen_proyectos_zonas = []
        for zona_nombre in todas_las_zonas:
            zona_dict = {'zona_operacional': zona_nombre}
            for proyecto in proyectos:
                zona_dict[proyecto] = resumen_por_proyecto[proyecto].get(zona_nombre, 0)
            resumen_proyectos_zonas.append(zona_dict)

        # Calcular totales generales para Totales por Zona Operacional
        total_general_especialidad = sum(z['total_especialidad'] for z in resumen_por_zona.values())
        total_general_vehiculo = sum(z['total_vehiculo'] for z in resumen_por_zona.values())

        # Calcular totales por proyecto (para la tabla proyectos vs zonas)
        totales_por_proyecto = {proyecto: 0 for proyecto in proyectos}
        for row in resumen_proyectos_zonas:
            for proyecto in proyectos:
                totales_por_proyecto[proyecto] += row[proyecto]

        # Cálculo de porcentajes en función del total global
        total_global = resumen['total_pago'] or 1  # evitar división por cero

        # Porcentajes para especialidad y vehiculo
        porcentaje_especialidad = (total_general_especialidad / total_global * 100) if total_global != 0 else 0
        porcentaje_vehiculo = (total_general_vehiculo / total_global * 100) if total_global != 0 else 0

        # Porcentajes por proyecto
        porcentajes_por_proyecto = {}
        for proyecto, valor in totales_por_proyecto.items():
            porcentajes_por_proyecto[proyecto] = (valor / total_global * 100) if total_global != 0 else 0

        return render(request, 'pagos/ver_resumen_mensual.html', {
            'resumen': resumen,
            'resumen_por_zona': resumen_por_zona,
            'resumen_proyectos_zonas': resumen_proyectos_zonas,
            'resumen_areas_final': resumen_areas_final,
            'total_general_area': total_general_area,
            'month': month,
            'year': year,
            'proyectos': proyectos,
            'total_general_especialidad': total_general_especialidad,
            'total_general_vehiculo': total_general_vehiculo,
            'totales_por_proyecto': totales_por_proyecto,
            'porcentaje_especialidad': porcentaje_especialidad,
            'porcentaje_vehiculo': porcentaje_vehiculo,
            'porcentajes_por_proyecto': porcentajes_por_proyecto
        })

    except Exception as e:
        print(f"Error: {e}")
        return render(request, 'pagos/error.html', {'mensaje': 'Ocurrió un error inesperado.'})

# pagos/views.py
from django.shortcuts import render, get_object_or_404, redirect
from .models import TablaDePagoCapacity
  # Ajusta según el nombre de tu modelo

def editar_registro(request, id):
    registro = get_object_or_404(TablaDePagoCapacity, id=id)
    if request.method == 'POST':
        # Aquí manejas el formulario de edición, validas datos, guardas cambios, etc.
        # Por ejemplo:
        registro.campo = request.POST.get('campo', registro.campo)
        registro.save()
        return redirect('pagos/tabla_pago_capacity_mes.html')

    # Si es GET, muestras la plantilla con un formulario de edición
    return render(request, 'pagos/editar_registro.html', {'registro': registro})


from django.shortcuts import get_object_or_404, redirect
from .models import TablaDePagoCapacity  # Ajusta el nombre del modelo según tu proyecto

def eliminar_registro(request, id):
    registro = get_object_or_404(TablaDePagoCapacity, id=id)
    if request.method == 'POST':
        # Si el método es POST, significa que el usuario ya confirmó la eliminación
        registro.delete()
        # Redirige a la vista que muestra la lista completa luego de eliminar
        return redirect('nombre_de_la_vista_de_listado')  
    # Si es GET, podría mostrar una plantilla de confirmación de eliminación
    # o podrías simplemente eliminar directamente.
    # Aquí solo mostramos un ejemplo de confirmación simple:
    return render(request, 'pagos/confirmar_eliminar.html', {'registro': registro})
